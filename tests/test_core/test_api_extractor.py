"""Tests for core.api_extractor module."""

import json
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock, mock_open

import pytest
from urllib.error import HTTPError, URLError

from verificacion_correo.core.api_extractor import (
    SessionExpiredError,
    find_people,
    process_emails_via_api,
    _build_cookie_header,
    _get_canary,
    _build_headers,
    _find_persona_id,
    _get_persona,
    _parse_persona,
    _extract_first_array_value,
    OWA_BASE,
    REQUEST_DELAY,
)
from verificacion_correo.core.extractor import ContactInfo
from verificacion_correo.core.excel import ProcessingStatus


SESSION_DATA = {
    "cookies": [
        {"name": "X-OWA-CANARY", "value": "test_canary_123", "domain": ".madrid.org"},
        {"name": "session_id", "value": "abc123", "domain": ".madrid.org"},
    ],
    "origins": [
        {
            "origin": "https://correoweb.madrid.org",
            "localStorage": [{"canary_key": "local_canary"}],
        }
    ],
}


class TestSessionExpiredError:
    def test_is_exception(self):
        assert issubclass(SessionExpiredError, Exception)

    def test_with_message(self):
        error = SessionExpiredError("Session expired")
        assert str(error) == "Session expired"

    def test_raise(self):
        with pytest.raises(SessionExpiredError):
            raise SessionExpiredError("Test")


class TestBuildCookieHeader:
    def test_success(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        result = _build_cookie_header(str(session_file))
        assert "X-OWA-CANARY=test_canary_123" in result
        assert "session_id=abc123" in result

    def test_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            _build_cookie_header("/nonexistent/state.json")

    def test_filters_by_domain(self, tmp_path):
        data = {
            "cookies": [
                {"name": "valid_cookie", "value": "valid", "domain": ".madrid.org"},
                {"name": "other_cookie", "value": "other", "domain": ".example.com"},
            ]
        }
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        result = _build_cookie_header(str(session_file))
        assert "valid_cookie=valid" in result
        assert "other_cookie" not in result

    def test_empty_cookies(self, tmp_path):
        data = {"cookies": []}
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        result = _build_cookie_header(str(session_file))
        assert result == ""

    def test_missing_cookies_key(self, tmp_path):
        data = {"other": "data"}
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        result = _build_cookie_header(str(session_file))
        assert result == ""


class TestGetCanary:
    def test_from_cookies(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        result = _get_canary(str(session_file))
        assert result == "test_canary_123"

    def test_from_local_storage(self, tmp_path):
        data = {
            "cookies": [],
            "origins": [
                {
                    "origin": "https://correoweb.madrid.org",
                    "localStorage": [{"my_canary_value": "ls_canary"}],
                }
            ],
        }
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        result = _get_canary(str(session_file))
        assert result == "ls_canary"

    def test_not_found(self, tmp_path):
        data = {"cookies": []}
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        result = _get_canary(str(session_file))
        assert result == ""

    def test_file_not_found(self):
        result = _get_canary("/nonexistent/state.json")
        assert result == ""


class TestBuildHeaders:
    def test_contains_required_keys(self):
        headers = _build_headers("canary123", "cookie_str", "FindPeople")
        assert headers["Action"] == "FindPeople"
        assert headers["X-OWA-CANARY"] == "canary123"
        assert headers["Cookie"] == "cookie_str"
        assert "Content-Type" in headers
        assert "User-Agent" in headers

    def test_different_action(self):
        headers = _build_headers("c", "c", "GetPersona")
        assert headers["Action"] == "GetPersona"
        assert headers["X-OWA-ActionName"] == "BrowseInDirectory"


class TestExtractFirstArrayValue:
    def test_valid_array(self):
        persona = {"Phones": [{"Value": "912345678"}]}
        result = _extract_first_array_value(persona, "Phones")
        assert result == "912345678"

    def test_empty_array(self):
        result = _extract_first_array_value({"Phones": []}, "Phones")
        assert result is None

    def test_missing_key(self):
        result = _extract_first_array_value({}, "Phones")
        assert result is None

    def test_not_list(self):
        result = _extract_first_array_value({"Phones": "string"}, "Phones")
        assert result is None

    def test_custom_value_key(self):
        persona = {"Items": [{"CustomKey": "test_value"}]}
        result = _extract_first_array_value(persona, "Items", value_key="CustomKey")
        assert result == "test_value"

    def test_non_dict_first_item(self):
        persona = {"Items": ["string_value"]}
        result = _extract_first_array_value(persona, "Items")
        assert result == "string_value"

    def test_none_item(self):
        persona = {"Items": [None]}
        result = _extract_first_array_value(persona, "Items")
        assert result is None

    def test_empty_value_string(self):
        persona = {"Items": [{"Value": ""}]}
        result = _extract_first_array_value(persona, "Items")
        assert result is None


class TestParsePersona:
    def test_full_persona(self):
        persona = {
            "DisplayName": "Juan Perez",
            "EmailAddress": {"EmailAddress": "juan.perez@madrid.org"},
            "ImAddress": "sip:juan.perez@madrid.org",
            "CompanyName": "Madrid Org",
            "Department": "IT",
            "OfficeLocation": "Edificio A",
            "BusinessPhoneNumbersArray": [{"Value": {"Number": "912345678"}}],
            "BusinessAddressesArray": [
                {
                    "Value": {
                        "Street": "C/ Mayor 1",
                        "City": "Madrid",
                        "State": "Madrid",
                        "PostalCode": "28001",
                        "Country": "Spain",
                    }
                }
            ],
        }
        contact = _parse_persona(persona)
        assert contact is not None
        assert contact.name == "Juan Perez"
        assert contact.email == "juan.perez@madrid.org"
        assert contact.phone == "912345678"
        assert contact.sip == "sip:juan.perez@madrid.org"
        assert "C/ Mayor 1" in contact.address
        assert "Madrid" in contact.address
        assert contact.department == "IT"
        assert contact.company == "Madrid Org"
        assert contact.office_location == "Edificio A"

    def test_minimal_persona(self):
        persona = {
            "DisplayName": "Test User",
            "EmailAddress": "test@madrid.org",
        }
        contact = _parse_persona(persona)
        assert contact is not None
        assert contact.name == "Test User"
        assert contact.email == "test@madrid.org"

    def test_persona_with_email_object(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": {"EmailAddress": "obj@madrid.org"},
        }
        contact = _parse_persona(persona)
        assert contact is not None
        assert contact.email == "obj@madrid.org"

    def test_persona_with_email_string(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "str@madrid.org",
        }
        contact = _parse_persona(persona)
        assert contact is not None
        assert contact.email == "str@madrid.org"

    def test_persona_no_valid_data_returns_none(self):
        persona = {"DisplayName": "Only Name"}
        contact = _parse_persona(persona)
        assert contact is None

    def test_persona_empty(self):
        contact = _parse_persona({})
        assert contact is None

    def test_phone_from_dict(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "test@madrid.org",
            "BusinessPhoneNumbersArray": [{"Value": {"NormalizedNumber": "+34912345678"}}],
        }
        contact = _parse_persona(persona)
        assert contact.phone == "+34912345678"

    def test_phone_from_string(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "test@madrid.org",
            "BusinessPhoneNumbersArray": [{"Value": "912345678"}],
        }
        contact = _parse_persona(persona)
        assert contact.phone == "912345678"

    def test_office_locations_array(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "test@madrid.org",
            "OfficeLocationsArray": [{"Value": "Main Office"}],
        }
        contact = _parse_persona(persona)
        assert contact.office_location == "Main Office"

    def test_office_location_fallback(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "test@madrid.org",
            "OfficeLocation": "Fallback Office",
        }
        contact = _parse_persona(persona)
        assert contact.office_location == "Fallback Office"

    def test_address_from_string(self):
        persona = {
            "DisplayName": "Test",
            "EmailAddress": "test@madrid.org",
            "BusinessAddressesArray": [{"Value": "123 Main St"}],
        }
        contact = _parse_persona(persona)
        assert contact.address == "123 Main St"


class TestFindPersonaId:
    def test_success(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {
                "ResponseCode": "NoError",
                "ResultSet": [
                    {"PersonaId": {"Id": "persona_123"}}
                ],
            }
        }).encode("utf-8")

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id == "persona_123"

    def test_persona_not_found(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {"ResponseCode": "ErrorNoResult"}
        }).encode("utf-8")

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id is None

    def test_empty_result_set(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {"ResponseCode": "NoError", "ResultSet": []}
        }).encode("utf-8")

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id is None

    def test_http_307_raises_session_expired(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        http_error = HTTPError(
            url="http://example.com",
            code=307,
            msg="Temporary Redirect",
            hdrs={},
            fp=None,
        )

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=http_error):
            with pytest.raises(SessionExpiredError):
                _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")

    def test_http_other_error_returns_none(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        http_error = HTTPError(
            url="http://example.com",
            code=404,
            msg="Not Found",
            hdrs={},
            fp=None,
        )

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=http_error):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id is None

    def test_url_error_timeout(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        url_error = URLError(reason=TimeoutError())

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=url_error):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id is None

    def test_url_error_connection(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        url_error = URLError(reason="Connection refused")

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=url_error):
            persona_id = _find_persona_id("test@madrid.org", cookie_str, canary, "addr_list_id")
            assert persona_id is None


class TestGetPersona:
    def test_success(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {
                "Persona": {
                    "DisplayName": "Juan Perez",
                    "EmailAddress": {"EmailAddress": "juan.perez@madrid.org"},
                }
            }
        }).encode("utf-8")

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            contact = _get_persona("persona_123", cookie_str, canary)
            assert contact is not None
            assert contact.name == "Juan Perez"
            assert contact.email == "juan.perez@madrid.org"

    def test_no_persona_key(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {}
        }).encode("utf-8")

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            contact = _get_persona("persona_123", cookie_str, canary)
            assert contact is None

    def test_http_307_raises_session_expired(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        http_error = HTTPError(
            url="http://example.com",
            code=307,
            msg="Temporary Redirect",
            hdrs={},
            fp=None,
        )

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=http_error):
            with pytest.raises(SessionExpiredError):
                _get_persona("persona_123", cookie_str, canary)

    def test_http_other_error(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        http_error = HTTPError(
            url="http://example.com",
            code=500,
            msg="Server Error",
            hdrs={},
            fp=None,
        )

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=http_error):
            contact = _get_persona("persona_123", cookie_str, canary)
            assert contact is None

    def test_url_error(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        cookie_str = _build_cookie_header(str(session_file))
        canary = _get_canary(str(session_file))

        url_error = URLError(reason="Connection error")

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=url_error):
            contact = _get_persona("persona_123", cookie_str, canary)
            assert contact is None


class TestFindPeople:
    def test_successful_lookup(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_find_response = MagicMock()
        mock_find_response.read.return_value = json.dumps({
            "Body": {
                "ResponseCode": "NoError",
                "ResultSet": [{"PersonaId": {"Id": "persona_123"}}],
            }
        }).encode("utf-8")

        mock_get_response = MagicMock()
        mock_get_response.read.return_value = json.dumps({
            "Body": {
                "Persona": {
                    "DisplayName": "Juan Perez",
                    "EmailAddress": {"EmailAddress": "juan.perez@madrid.org"},
                }
            }
        }).encode("utf-8")

        with patch(
            "verificacion_correo.core.api_extractor.urlopen",
            side_effect=[mock_find_response, mock_get_response],
        ):
            contact = find_people("test@madrid.org", str(session_file))
            assert contact is not None
            assert contact.name == "Juan Perez"

    def test_no_persona_id(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_response = MagicMock()
        mock_response.read.return_value = json.dumps({
            "Body": {"ResponseCode": "ErrorNoResult"}
        }).encode("utf-8")

        with patch("verificacion_correo.core.api_extractor.urlopen", return_value=mock_response):
            contact = find_people("test@madrid.org", str(session_file))
            assert contact is None

    def test_get_persona_returns_none(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        mock_find_response = MagicMock()
        mock_find_response.read.return_value = json.dumps({
            "Body": {
                "ResponseCode": "NoError",
                "ResultSet": [{"PersonaId": {"Id": "persona_123"}}],
            }
        }).encode("utf-8")

        mock_get_response = MagicMock()
        mock_get_response.read.return_value = json.dumps({
            "Body": {}
        }).encode("utf-8")

        with patch(
            "verificacion_correo.core.api_extractor.urlopen",
            side_effect=[mock_find_response, mock_get_response],
        ):
            contact = find_people("test@madrid.org", str(session_file))
            assert contact is None

    def test_session_expired_during_find(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        http_error = HTTPError(
            url="http://example.com",
            code=307,
            msg="Redirect",
            hdrs={},
            fp=None,
        )

        with patch("verificacion_correo.core.api_extractor.urlopen", side_effect=http_error):
            with pytest.raises(SessionExpiredError):
                find_people("test@madrid.org", str(session_file))

    def test_no_canary_returns_none(self, tmp_path):
        data = {"cookies": []}
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(data, f)

        contact = find_people("test@madrid.org", str(session_file))
        assert contact is None


class TestProcessEmailsViaApi:
    def test_no_pending_emails(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "empty.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="done@test.org")
        ws.cell(row=2, column=2, value="OK")
        wb.save(str(excel_path))
        wb.close()

        stats = process_emails_via_api(str(excel_path), str(session_file))
        assert stats["total"] == 0
        assert stats["success"] == 0
        assert stats["expired"] is False

    def test_successful_batch(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="pending@test.org")
        ws.cell(row=2, column=2, value="")
        wb.save(str(excel_path))
        wb.close()

        with patch("verificacion_correo.core.api_extractor.find_people") as mock_find:
            mock_find.return_value = ContactInfo(
                name="Test User",
                email="test@madrid.org",
            )
            stats = process_emails_via_api(str(excel_path), str(session_file))

            assert stats["success"] == 1
            assert stats["not_found"] == 0
            assert stats["errors"] == 0
            assert stats["total"] == 1

    def test_not_found_in_batch(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="missing@test.org")
        ws.cell(row=2, column=2, value="")
        wb.save(str(excel_path))
        wb.close()

        with patch("verificacion_correo.core.api_extractor.find_people", return_value=None):
            stats = process_emails_via_api(str(excel_path), str(session_file))
            assert stats["not_found"] == 1
            assert stats["success"] == 0
            assert stats["total"] == 1

    def test_session_expired_during_processing(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="first@test.org")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=3, column=1, value="second@test.org")
        ws.cell(row=3, column=2, value="")
        wb.save(str(excel_path))
        wb.close()

        with patch("verificacion_correo.core.api_extractor.find_people") as mock_find:
            mock_find.side_effect = SessionExpiredError("Session expired")

            stats = process_emails_via_api(str(excel_path), str(session_file))
            assert stats["errors"] == 1
            assert stats["total"] == 1
            assert stats["expired"] is True
            assert stats["remaining"] == 1

    def test_progress_callback(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="a@test.org")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=3, column=1, value="b@test.org")
        ws.cell(row=3, column=2, value="")
        wb.save(str(excel_path))
        wb.close()

        callback_calls = []

        with patch("verificacion_correo.core.api_extractor.find_people") as mock_find:
            mock_find.return_value = ContactInfo(email="found@madrid.org")

            stats = process_emails_via_api(
                str(excel_path), str(session_file),
                progress_callback=lambda p, t: callback_calls.append((p, t)),
            )
            assert stats["total"] == 2
            assert len(callback_calls) == 2
            assert callback_calls[0] == (1, 2)
            assert callback_calls[1] == (2, 2)

    def test_custom_address_list_id(self, tmp_path):
        session_file = tmp_path / "state.json"
        with open(session_file, "w") as f:
            json.dump(SESSION_DATA, f)

        excel_path = tmp_path / "test.xlsx"
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Correo")
        ws.cell(row=1, column=2, value="Status")
        ws.cell(row=2, column=1, value="test@test.org")
        ws.cell(row=2, column=2, value="")
        wb.save(str(excel_path))
        wb.close()

        with patch("verificacion_correo.core.api_extractor.find_people") as mock_find:
            mock_find.return_value = ContactInfo(email="test@madrid.org")

            stats = process_emails_via_api(
                str(excel_path), str(session_file),
                address_list_id="custom_list_id",
            )
            assert stats["success"] == 1
            mock_find.assert_called_with("test@test.org", str(session_file), "custom_list_id")


class TestConstants:
    def test_owa_base(self):
        assert OWA_BASE == "https://correoweb.madrid.org"

    def test_request_delay(self):
        assert REQUEST_DELAY == 3.0
