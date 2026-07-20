"""Tests para verificacion_correo.core.gal_enricher."""

import pytest
from pathlib import Path
from openpyxl import Workbook
from verificacion_correo.core.gal_enricher import (
    EnrichProgress,
    merge_enrichment,
    get_companies_to_enrich_from_excel,
)


class TestEnrichProgress:
    """Tests para EnrichProgress."""

    def test_save_and_load(self, tmp_path):
        progress = EnrichProgress(tmp_path / "progress.json")
        progress.data['companies_done'] = ['A', 'B']
        progress.data['contacts_enriched'] = 50
        progress.save()

        progress2 = EnrichProgress(tmp_path / "progress.json")
        assert progress2.load()
        assert progress2.data['companies_done'] == ['A', 'B']
        assert progress2.data['contacts_enriched'] == 50

    def test_load_returns_false_when_no_file(self, tmp_path):
        progress = EnrichProgress(tmp_path / "nonexistent.json")
        assert progress.load() is False


class TestMergeEnrichment:
    """Tests para merge_enrichment."""

    def test_merge_fills_empty_fields(self):
        existing = {'telefono': '', 'departamento': 'IT', 'oficina': '', 'direccion': ''}
        enrichment = {'telefono': '123456', 'departamento': '', 'oficina': 'Office A', 'direccion': 'Calle 1'}
        result = merge_enrichment(existing, enrichment)
        assert result['telefono'] == '123456'
        assert result['departamento'] == 'IT'
        assert result['oficina'] == 'Office A'
        assert result['direccion'] == 'Calle 1'

    def test_merge_does_not_overwrite_existing(self):
        existing = {'telefono': '111', 'departamento': 'IT'}
        enrichment = {'telefono': '222', 'departamento': 'Sales'}
        result = merge_enrichment(existing, enrichment)
        assert result['telefono'] == '111'
        assert result['departamento'] == 'IT'

    def test_merge_ignores_none_values(self):
        existing = {'telefono': ''}
        enrichment = {'telefono': None}
        result = merge_enrichment(existing, enrichment)
        assert result['telefono'] == ''


class TestGetCompaniesFromExcel:
    """Tests para get_companies_to_enrich_from_excel."""

    def test_returns_marked_companies(self, tmp_path):
        wb = Workbook()
        ws2 = wb.create_sheet("Compañías")
        ws2.append(['Compañía', 'Enrich'])
        ws2.append(['ORG1', 'X'])
        ws2.append(['ORG2', ''])
        ws2.append(['ORG3', 'X'])

        excel_path = tmp_path / "test.xlsx"
        wb.save(excel_path)

        result = get_companies_to_enrich_from_excel(excel_path)
        assert result == ['ORG1', 'ORG3']

    def test_empty_when_no_x(self, tmp_path):
        wb = Workbook()
        ws2 = wb.create_sheet("Compañías")
        ws2.append(['Compañía', 'Enrich'])
        ws2.append(['ORG1', ''])

        excel_path = tmp_path / "test.xlsx"
        wb.save(excel_path)

        result = get_companies_to_enrich_from_excel(excel_path)
        assert result == []
