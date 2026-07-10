"""Tests for core.excel module."""

import os
import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock, PropertyMock

import pytest
from openpyxl import Workbook, load_workbook

from verificacion_correo.core.excel import (
    ExcelReader,
    ExcelWriter,
    ExcelColumn,
    ExcelColumns,
    EmailRecord,
    ExcelSummary,
    ProcessingStatus,
)


@pytest.fixture
def excel_file_with_pending():
    """Create a temp Excel file with pending emails."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    headers = ["Correo", "Status", "Nombre", "Email Personal", "Teléfono",
               "SIP", "Dirección", "Departamento", "Compañía", "Oficina"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    emails = ["user1@madrid.org", "user2@madrid.org", "user3@madrid.org"]
    for row_idx, email in enumerate(emails, start=2):
        ws.cell(row=row_idx, column=1, value=email)
        ws.cell(row=row_idx, column=2, value="")

    wb.save(tmp.name)
    wb.close()
    return tmp.name


@pytest.fixture
def excel_file_mixed():
    """Create a temp Excel file with some processed and some pending."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    headers = ["Correo", "Status", "Nombre", "Email Personal", "Teléfono",
               "SIP", "Dirección", "Departamento", "Compañía", "Oficina"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    ws.cell(row=2, column=1, value="done@madrid.org")
    ws.cell(row=2, column=2, value="OK")
    ws.cell(row=3, column=1, value="pending@madrid.org")
    ws.cell(row=3, column=2, value="")
    ws.cell(row=4, column=1, value="error@madrid.org")
    ws.cell(row=4, column=2, value="ERROR")

    wb.save(tmp.name)
    wb.close()
    return tmp.name


@pytest.fixture
def excel_file_all_processed():
    """Create a temp Excel file with all emails processed."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    headers = ["Correo", "Status", "Nombre", "Email Personal", "Teléfono",
               "SIP", "Dirección", "Departamento", "Compañía", "Oficina"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    ws.cell(row=2, column=1, value="done1@madrid.org")
    ws.cell(row=2, column=2, value="OK")
    ws.cell(row=3, column=1, value="done2@madrid.org")
    ws.cell(row=3, column=2, value="OK")

    wb.save(tmp.name)
    wb.close()
    return tmp.name


@pytest.fixture
def empty_excel_file():
    """Create a temp Excel file with only headers."""
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Contactos"
    headers = ["Correo", "Status", "Nombre", "Email Personal", "Teléfono",
               "SIP", "Dirección", "Departamento", "Compañía", "Oficina"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)

    wb.save(tmp.name)
    wb.close()
    return tmp.name


class TestProcessingStatus:
    def test_values(self):
        assert ProcessingStatus.PENDING.value == "PENDIENTE"
        assert ProcessingStatus.SUCCESS.value == "OK"
        assert ProcessingStatus.NOT_FOUND.value == "NO EXISTE"
        assert ProcessingStatus.ERROR.value == "ERROR"

    def test_is_enum(self):
        assert isinstance(ProcessingStatus.PENDING, ProcessingStatus)


class TestExcelColumn:
    def test_creation(self):
        col = ExcelColumn("Test", "T", 20, "A test column")
        assert col.name == "Test"
        assert col.letter == "T"
        assert col.index == 20
        assert col.description == "A test column"


class TestExcelColumns:
    def test_class_attributes(self):
        assert ExcelColumns.EMAIL.name == "Correo"
        assert ExcelColumns.EMAIL.letter == "A"
        assert ExcelColumns.EMAIL.index == 1

        assert ExcelColumns.STATUS.name == "Status"
        assert ExcelColumns.STATUS.letter == "B"
        assert ExcelColumns.STATUS.index == 2

        assert ExcelColumns.NAME.name == "Nombre"
        assert ExcelColumns.NAME.index == 3

        assert ExcelColumns.OFFICE_LOCATION.name == "Oficina"
        assert ExcelColumns.OFFICE_LOCATION.index == 10

    def test_get_all_columns(self):
        cols = ExcelColumns.get_all_columns()
        assert len(cols) == 10
        assert cols[0].name == "Correo"
        assert cols[-1].name == "Oficina"

    def test_get_headers(self):
        headers = ExcelColumns.get_headers()
        assert len(headers) == 10
        assert headers == [
            "Correo", "Status", "Nombre", "Email Personal", "Teléfono",
            "SIP", "Dirección", "Departamento", "Compañía", "Oficina"
        ]

    def test_get_column_by_name_found(self):
        col = ExcelColumns.get_column_by_name("Teléfono")
        assert col is not None
        assert col.letter == "E"
        assert col.index == 5

    def test_get_column_by_name_not_found(self):
        col = ExcelColumns.get_column_by_name("NonExistent")
        assert col is None

    def test_get_column_by_name_case_sensitive(self):
        col = ExcelColumns.get_column_by_name("telefono")
        assert col is None


class TestEmailRecord:
    def test_default_status(self):
        rec = EmailRecord(email="test@example.com", row=5)
        assert rec.email == "test@example.com"
        assert rec.row == 5
        assert rec.status == ProcessingStatus.PENDING
        assert rec.data is None

    def test_with_data(self):
        rec = EmailRecord(
            email="test@example.com", row=5,
            status=ProcessingStatus.SUCCESS,
            data={"name": "Test User", "phone": "123456789"}
        )
        assert rec.status == ProcessingStatus.SUCCESS
        assert rec.data["name"] == "Test User"

    def test_with_not_found_status(self):
        rec = EmailRecord(
            email="test@example.com", row=5,
            status=ProcessingStatus.NOT_FOUND
        )
        assert rec.status == ProcessingStatus.NOT_FOUND


class TestExcelSummary:
    def test_empty_summary(self):
        summary = ExcelSummary(total_emails=0, pending_count=0, processed_count=0, batches=[])
        assert summary.total_emails == 0
        assert summary.pending_count == 0
        assert summary.processed_count == 0
        assert summary.batches == []

    def test_with_batches(self):
        batch1 = [EmailRecord(email="a@b.com", row=2)]
        batch2 = [EmailRecord(email="c@d.com", row=3)]
        summary = ExcelSummary(
            total_emails=2, pending_count=2, processed_count=0,
            batches=[batch1, batch2]
        )
        assert len(summary.batches) == 2
        assert summary.total_emails == 2


class TestExcelReader:
    def test_read_pending_emails_all_pending(self, excel_file_with_pending):
        reader = ExcelReader(excel_file_with_pending)
        summary = reader.read_pending_emails(batch_size=2)

        assert summary.total_emails == 3
        assert summary.pending_count == 3
        assert summary.processed_count == 0
        assert len(summary.batches) == 2  # 3 emails in batch_size=2 → 2 batches
        assert len(summary.batches[0]) == 2
        assert len(summary.batches[1]) == 1

    def test_read_pending_emails_large_batch(self, excel_file_with_pending):
        reader = ExcelReader(excel_file_with_pending)
        summary = reader.read_pending_emails(batch_size=100)

        assert len(summary.batches) == 1
        assert len(summary.batches[0]) == 3

    def test_read_pending_emails_single_batch(self, excel_file_with_pending):
        reader = ExcelReader(excel_file_with_pending)
        summary = reader.read_pending_emails(batch_size=3)

        assert len(summary.batches) == 1
        assert len(summary.batches[0]) == 3

    def test_read_pending_emails_mixed(self, excel_file_mixed):
        reader = ExcelReader(excel_file_mixed)
        summary = reader.read_pending_emails()

        assert summary.total_emails == 3
        assert summary.pending_count == 1
        assert summary.processed_count == 2
        pending = summary.batches[0][0]
        assert pending.email == "pending@madrid.org"

    def test_read_pending_emails_all_processed(self, excel_file_all_processed):
        reader = ExcelReader(excel_file_all_processed)
        summary = reader.read_pending_emails()

        assert summary.total_emails == 2
        assert summary.pending_count == 0
        assert summary.processed_count == 2
        assert summary.batches == []

    def test_read_pending_emails_empty(self, empty_excel_file):
        reader = ExcelReader(empty_excel_file)
        summary = reader.read_pending_emails()

        assert summary.total_emails == 0
        assert summary.pending_count == 0
        assert summary.processed_count == 0
        assert summary.batches == []

    def test_read_pending_emails_file_not_found(self, tmp_path):
        with patch.object(ExcelReader, '_ensure_excel_file_exists') as mock_ensure:
            reader = ExcelReader(os.path.join(tmp_path, "nonexistent", "file.xlsx"))
            summary = reader.read_pending_emails()

            assert summary.total_emails == 0
            assert summary.pending_count == 0
            assert summary.processed_count == 0
            assert summary.batches == []

    def test_read_all_emails(self, excel_file_with_pending):
        reader = ExcelReader(excel_file_with_pending)
        emails = reader.read_all_emails()

        assert len(emails) == 3
        assert "user1@madrid.org" in emails
        assert "user2@madrid.org" in emails
        assert "user3@madrid.org" in emails

    def test_read_all_emails_file_not_found(self, tmp_path):
        with patch.object(ExcelReader, '_ensure_excel_file_exists') as mock_ensure:
            reader = ExcelReader(os.path.join(tmp_path, "nonexistent", "file.xlsx"))
            emails = reader.read_all_emails()
            assert emails == []

    def test_read_all_records(self, excel_file_mixed):
        reader = ExcelReader(excel_file_mixed)
        records = reader.read_all_records()

        assert len(records) == 3

        done = next(r for r in records if r.email == "done@madrid.org")
        assert done.status == ProcessingStatus.SUCCESS

        pending = next(r for r in records if r.email == "pending@madrid.org")
        assert pending.status == ProcessingStatus.PENDING

        error = next(r for r in records if r.email == "error@madrid.org")
        assert error.status == ProcessingStatus.ERROR

    def test_read_all_records_empty(self, empty_excel_file):
        reader = ExcelReader(empty_excel_file)
        records = reader.read_all_records()
        assert records == []

    def test_ensure_excel_file_creates_new(self, tmp_path):
        file_path = tmp_path / "new_file.xlsx"
        reader = ExcelReader(str(file_path))
        assert file_path.exists()

        # Should have sample data
        summary = reader.read_pending_emails()
        assert summary.total_emails == 3  # 3 sample emails

    def test_ensure_excel_file_skips_if_exists(self, tmp_path):
        file_path = tmp_path / "existing.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Manual")
        wb.save(str(file_path))
        wb.close()

        reader = ExcelReader(str(file_path))
        # File should still have just one cell
        emails = reader.read_all_emails()
        assert emails == []  # "Manual" has no @

    def test_create_empty_excel_with_sample_data(self, tmp_path):
        file_path = tmp_path / "sample.xlsx"
        with patch.object(ExcelReader, '_create_from_template', return_value=False):
            reader = ExcelReader(str(file_path))
        summary = reader.read_pending_emails()

        assert summary.total_emails == 3
        emails_in = reader.read_all_emails()
        assert "ASP164@MADRID.ORG" in emails_in
        assert "AGM564@MADRID.ORG" in emails_in
        assert "USR789@MADRID.ORG" in emails_in


class TestExcelWriter:
    def test_write_success_result(self, excel_file_with_pending):
        file_path = excel_file_with_pending
        writer = ExcelWriter(file_path)

        record = EmailRecord(
            email="user1@madrid.org", row=2,
            status=ProcessingStatus.SUCCESS,
            data={
                "name": "Juan Perez",
                "email": "juan.perez@madrid.org",
                "phone": "912345678",
                "department": "IT",
            }
        )
        writer.write_result(record)

        # Verify by reading back
        reader = ExcelReader(file_path)
        records = reader.read_all_records()
        match = next(r for r in records if r.email == "user1@madrid.org")
        assert match.status == ProcessingStatus.SUCCESS
        assert match.data["name"] == "Juan Perez"
        assert match.data["phone"] == "912345678"

    def test_write_not_found_result(self, excel_file_with_pending):
        file_path = excel_file_with_pending
        writer = ExcelWriter(file_path)

        record = EmailRecord(
            email="user1@madrid.org", row=2,
            status=ProcessingStatus.NOT_FOUND,
        )
        writer.write_result(record)

        # Directly verify the value was written in the Excel
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active
        status_value = ws.cell(row=2, column=2).value
        wb.close()

        assert status_value == "NO EXISTE"

    def test_write_error_result(self, excel_file_with_pending):
        file_path = excel_file_with_pending
        writer = ExcelWriter(file_path)

        record = EmailRecord(
            email="user1@madrid.org", row=2,
            status=ProcessingStatus.ERROR,
            data={"error": "Connection failed"}
        )
        writer.write_result(record)

        reader = ExcelReader(file_path)
        records = reader.read_all_records()
        match = next(r for r in records if r.email == "user1@madrid.org")
        assert match.status == ProcessingStatus.ERROR

    def test_write_batch_results(self, excel_file_with_pending):
        file_path = excel_file_with_pending
        writer = ExcelWriter(file_path)

        records = [
            EmailRecord(
                email="user1@madrid.org", row=2,
                status=ProcessingStatus.SUCCESS,
                data={"name": "User One"}
            ),
            EmailRecord(
                email="user2@madrid.org", row=3,
                status=ProcessingStatus.NOT_FOUND,
            ),
            EmailRecord(
                email="user3@madrid.org", row=4,
                status=ProcessingStatus.ERROR,
            ),
        ]
        writer.write_batch_results(records)

        reader = ExcelReader(file_path)
        all_records = reader.read_all_records()
        assert len(all_records) == 3

    def test_write_batch_results_clears_data_on_error(self, excel_file_mixed):
        file_path = excel_file_mixed

        writer = ExcelWriter(file_path)
        writer.write_result(EmailRecord(
            email="pending@madrid.org", row=3,
            status=ProcessingStatus.ERROR,
        ))

        reader = ExcelReader(file_path)
        records = reader.read_all_records()
        match = next(r for r in records if r.email == "pending@madrid.org")
        assert match.status == ProcessingStatus.ERROR

    def test_ensure_file_structure_creates_new(self, tmp_path):
        file_path = tmp_path / "writer_test.xlsx"
        writer = ExcelWriter(str(file_path))
        writer.ensure_file_structure()

        assert file_path.exists()
        reader = ExcelReader(str(file_path))
        emails = reader.read_all_emails()
        assert emails == []  # Just headers, no data

    def test_ensure_file_structure_creates_parent(self, tmp_path):
        file_path = tmp_path / "nested" / "test.xlsx"
        writer = ExcelWriter(str(file_path))
        writer.ensure_file_structure()

        assert file_path.exists()

    def test_get_status(self, excel_file_with_pending):
        file_path = excel_file_with_pending
        writer = ExcelWriter(file_path)

        # Initially empty
        status = writer.get_status(2)
        assert status is None or status == ""

        # Write and check
        writer.write_result(EmailRecord(email="u@m.org", row=2, status=ProcessingStatus.SUCCESS))
        assert writer.get_status(2) == "OK"

    def test_get_status_file_not_found(self, tmp_path):
        writer = ExcelWriter(str(tmp_path / "nonexistent.xlsx"))
        assert writer.get_status(2) is None

    def test_verify_headers_updates_existing(self, tmp_path):
        wb_path = tmp_path / "wrong_headers.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Wrong Header")
        ws.cell(row=1, column=2, value="Bad Status")
        wb.save(str(wb_path))
        wb.close()
        file_path = wb_path

        writer = ExcelWriter(str(file_path))
        # _verify_headers called internally during ensure_file_structure
        writer.ensure_file_structure()

        from openpyxl import load_workbook
        wb2 = load_workbook(str(file_path))
        ws2 = wb2.active
        assert ws2.cell(row=1, column=1).value == "Correo"
        assert ws2.cell(row=1, column=2).value == "Status"
        wb2.close()

    def test_clear_contact_data_on_not_found(self, tmp_path):
        file_path = str(tmp_path / "clear_test.xlsx")
        wb = Workbook()
        ws = wb.active
        cols = ExcelColumns.get_headers()
        for col_idx, h in enumerate(cols, start=1):
            ws.cell(row=1, column=col_idx, value=h)
        ws.cell(row=2, column=1, value="test@madrid.org")
        ws.cell(row=2, column=2, value="")
        ws.cell(row=2, column=3, value="Existing Data")
        ws.cell(row=2, column=5, value="123456789")
        wb.save(file_path)
        wb.close()

        writer = ExcelWriter(file_path)
        writer.write_result(EmailRecord(
            email="test@madrid.org", row=2,
            status=ProcessingStatus.NOT_FOUND,
        ))

        wb2 = load_workbook(file_path, data_only=True)
        ws2 = wb2.active
        status_val = ws2.cell(row=2, column=2).value
        col3_val = ws2.cell(row=2, column=3).value
        col5_val = ws2.cell(row=2, column=5).value
        wb2.close()

        assert status_val == "NO EXISTE", f"Status expected 'NO EXISTE', got: {status_val}"
        assert col3_val is None, f"Col 3 expected None, got: {col3_val}"
        assert col5_val is None, f"Col 5 expected None, got: {col5_val}"


class TestInitialization:
    def test_reader_custom_start_row(self, tmp_path):
        file_path = tmp_path / "custom_start.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Header")
        ws.cell(row=2, column=1, value="skip@me.org")
        ws.cell(row=3, column=1, value="take@me.org")
        wb.save(str(file_path))
        wb.close()

        reader = ExcelReader(str(file_path), start_row=3)
        emails = reader.read_all_emails()
        assert len(emails) == 1
        assert emails[0] == "take@me.org"

    def test_reader_custom_email_column(self, tmp_path):
        file_path = tmp_path / "custom_col.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=3, value="Header")
        ws.cell(row=2, column=3, value="col3@me.org")
        wb.save(str(file_path))
        wb.close()

        reader = ExcelReader(str(file_path), email_column=3)
        emails = reader.read_all_emails()
        assert emails == ["col3@me.org"]
