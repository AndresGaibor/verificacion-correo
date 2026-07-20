"""Tests para verificacion_correo.core.gal_exporter."""

import pytest
from pathlib import Path
from openpyxl import load_workbook
from verificacion_correo.core.gal_exporter import (
    flatten_contact_to_dict,
    append_contacts_to_excel,
    load_gal_from_excel,
)


class TestFlattenContact:
    """Tests para flatten_contact_to_dict."""

    def test_flatten_contact_basic(self):
        persona = {
            'DisplayName': 'Juan Perez',
            'EmailAddresses': [{'Value': 'juan@madrid.org'}],
            'CompanyName': 'ORGANOS JUDICIALES',
        }
        result = flatten_contact_to_dict(persona)
        assert result['nombre'] == 'Juan Perez'
        assert result['email'] == 'juan@madrid.org'
        assert result['empresa'] == 'ORGANOS JUDICIALES'

    def test_flatten_contact_no_email(self):
        persona = {
            'DisplayName': 'Test User',
            'CompanyName': 'TEST CO',
        }
        result = flatten_contact_to_dict(persona)
        assert result['email'] == ''
        assert result['empresa'] == 'TEST CO'

    def test_flatten_contact_empty_fields(self):
        persona = {}
        result = flatten_contact_to_dict(persona)
        assert result['nombre'] == ''
        assert result['email'] == ''
        assert result['empresa'] == ''


class TestAppendContactsToExcel:
    """Tests para append_contacts_to_excel (upsert por persona_id)."""

    def test_append_creates_excel_with_two_sheets(self, tmp_path):
        contacts = [
            {'nombre': 'Test', 'email': 'test@test.com', 'empresa': 'TEST CO',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'abc123'},
        ]
        excel_path = tmp_path / "test.xlsx"
        append_contacts_to_excel(contacts, excel_path)

        wb = load_workbook(excel_path)
        assert "Contactos" in wb.sheetnames
        assert "Compañías" in wb.sheetnames

    def test_append_contactos_headers(self, tmp_path):
        contacts = [
            {'nombre': 'Test', 'email': 'test@test.com', 'empresa': 'TEST CO',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'abc123'},
        ]
        excel_path = tmp_path / "test.xlsx"
        append_contacts_to_excel(contacts, excel_path)

        wb = load_workbook(excel_path)
        ws1 = wb["Contactos"]
        headers = [cell.value for cell in ws1[1]]
        assert headers == ['nombre', 'email', 'empresa', 'telefono', 'departamento', 'oficina', 'direccion', 'persona_id']

    def test_append_companias_sheet(self, tmp_path):
        contacts = [
            {'nombre': 'Test', 'email': 'test@test.com', 'empresa': 'TEST CO',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'abc123'},
        ]
        excel_path = tmp_path / "test.xlsx"
        append_contacts_to_excel(contacts, excel_path)

        wb = load_workbook(excel_path)
        ws2 = wb["Compañías"]
        assert ws2.cell(1, 1).value == 'Compañía'
        assert ws2.cell(1, 2).value == 'Enrich'
        assert ws2.cell(2, 1).value == 'TEST CO'

    def test_append_upserts_by_persona_id(self, tmp_path):
        """Cuando el mismo persona_id existe, se actualiza; si no, se inserta."""
        excel_path = tmp_path / "test.xlsx"

        contacts1 = [{
            'nombre': 'Juan', 'email': 'juan@test.com', 'empresa': 'Ayto',
            'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '',
            'persona_id': 'abc123'
        }]
        append_contacts_to_excel(contacts1, excel_path)

        wb = load_workbook(excel_path)
        ws1 = wb["Contactos"]
        assert ws1.max_row == 2

        contacts2 = [{
            'nombre': 'Juan', 'email': 'juan@test.com', 'empresa': 'Ayto',
            'telefono': '912345678', 'departamento': 'IT', 'oficina': 'Oficina 1',
            'direccion': 'Calle Mayor 1', 'persona_id': 'abc123'
        }]
        append_contacts_to_excel(contacts2, excel_path)

        wb = load_workbook(excel_path)
        ws1 = wb["Contactos"]
        assert ws1.max_row == 2
        assert ws1.cell(2, 4).value == '912345678'

    def test_append_multiple_companies(self, tmp_path):
        contacts = [
            {'nombre': 'A', 'email': 'a@test.com', 'empresa': 'COMP A',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'id1'},
            {'nombre': 'B', 'email': 'b@test.com', 'empresa': 'COMP B',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'id2'},
            {'nombre': 'C', 'email': 'c@test.com', 'empresa': 'COMP A',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'id3'},
        ]
        excel_path = tmp_path / "test.xlsx"
        append_contacts_to_excel(contacts, excel_path)

        wb = load_workbook(excel_path)
        ws2 = wb["Compañías"]
        companies = [ws2.cell(r, 1).value for r in range(2, ws2.max_row + 1)]
        assert companies == ['COMP A', 'COMP B']


class TestLoadGalFromExcel:
    """Tests para load_gal_from_excel."""

    def test_load_from_empty_excel(self, tmp_path):
        excel_path = tmp_path / "test.xlsx"
        result = load_gal_from_excel(excel_path)
        assert result == []

    def test_load_returns_contacts(self, tmp_path):
        contacts = [
            {'nombre': 'Test', 'email': 'test@test.com', 'empresa': 'TEST CO',
             'telefono': '', 'departamento': '', 'oficina': '', 'direccion': '', 'persona_id': 'abc123'},
        ]
        excel_path = tmp_path / "test.xlsx"
        append_contacts_to_excel(contacts, excel_path)

        loaded = load_gal_from_excel(excel_path)
        assert len(loaded) == 1
        assert loaded[0]['nombre'] == 'Test'
