"""
Tests de integración para flujos completos entre módulos.

Valida interacciones reales entre ProgressFile, scrape_gal,
ExcelReader y ExcelWriter con archivos temporales.
"""

import json
from pathlib import Path
from unittest.mock import patch, MagicMock

import pytest

from verificacion_correo.core.gal_scraper import (
    ProgressFile,
    scrape_gal,
    _flatten_persona,
    PROGRESS_FILENAME,
    OUTPUT_JSON,
    OUTPUT_CSV,
)
from verificacion_correo.core.excel import (
    ExcelReader,
    ExcelWriter,
    EmailRecord,
    ProcessingStatus,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _create_session_file(path: Path) -> None:
    """Create a minimal session file with cookies and canary."""
    state = {
        "cookies": [
            {"name": "X-OWA-CANARY", "value": "FAKE-CANARY-123", "domain": ".madrid.org"},
        ],
    }
    path.write_text(json.dumps(state), encoding="utf-8")


def _mock_find_people_response(people: list) -> bytes:
    """Build a fake FindPeople JSON response."""
    return json.dumps({"Body": {"People": people}}).encode("utf-8")


SAMPLE_PERSONA = {
    "DisplayName": "Juan Pérez",
    "EmailAddress": {"EmailAddress": "juan@madrid.org"},
    "BusinessPhoneNumbersArray": [
        {"Value": {"Number": "912345678"}},
    ],
    "ImAddress": "sip:juan@madrid.org",
    "CompanyName": "Ayuntamiento",
    "Department": "IT",
    "OfficeLocation": "Calle Mayor 1",
    "BusinessAddressesArray": [
        {"Value": {"Street": "Calle Mayor 1", "City": "Madrid", "PostalCode": "28001"}},
    ],
}


# ===========================================================================
# TestGALScrapingFlow
# ===========================================================================


class TestGALScrapingFlow:
    """Flujos de integración para ProgressFile y scrape_gal."""

    def test_progress_file_save_load_cycle(self, tmp_path: Path) -> None:
        """ProgressFile save → load preserva datos entre sesiones."""
        pf = ProgressFile(tmp_path)

        assert not pf.exists
        assert pf.load() == {"offset": 0, "people": [], "completed_companies": []}

        people = [SAMPLE_PERSONA]
        pf.save(offset=100, people=people)

        assert pf.exists

        loaded = pf.load()
        assert loaded["offset"] == 100
        assert loaded["count"] == 1
        assert loaded["last_update"]  # non-empty timestamp
        assert len(loaded["people"]) == 0  # save() does not persist people list

    def test_progress_file_clear(self, tmp_path: Path) -> None:
        """ProgressFile.clear elimina el archivo de progreso."""
        pf = ProgressFile(tmp_path)
        pf.save(offset=50, people=[SAMPLE_PERSONA])
        assert pf.exists

        pf.clear()
        assert not pf.exists
        assert pf.load() == {"offset": 0, "people": [], "completed_companies": []}

    def test_progress_file_load_missing_returns_defaults(self, tmp_path: Path) -> None:
        """load() en directorio vacío retorna defaults."""
        pf = ProgressFile(tmp_path / "nonexistent")
        result = pf.load()
        assert result == {"offset": 0, "people": [], "completed_companies": []}

    def test_progress_file_load_corrupt_json(self, tmp_path: Path) -> None:
        """load() con JSON corrupto levanta error."""
        pf = ProgressFile(tmp_path)
        pf.path.write_text("NOT JSON {{{", encoding="utf-8")
        with pytest.raises(json.JSONDecodeError):
            pf.load()

    def test_scrape_gal_resume_flow(self, tmp_path: Path) -> None:
        """scrape_gal reanuda correctamente desde progreso guardado."""
        session_file = tmp_path / "state.json"
        _create_session_file(session_file)
        output_dir = str(tmp_path / "output")

        # Pre-populate progress: simulate 150 contacts already scraped
        pf = ProgressFile(Path(output_dir))
        pf.save(offset=150, people=[SAMPLE_PERSONA] * 150)

        # Mock HTTP to return 2 contacts then empty (stop condition)
        response_people = [
            {"DisplayName": "María García", "EmailAddress": {"EmailAddress": "maria@madrid.org"}},
            {"DisplayName": "Pedro López", "EmailAddress": {"EmailAddress": "pedro@madrid.org"}},
        ]

        call_count = 0

        def mock_urlopen(req, timeout=60):
            nonlocal call_count
            call_count += 1
            if call_count == 1:
                resp = MagicMock()
                resp.read.return_value = _mock_find_people_response(response_people)
                resp.__enter__ = lambda s: s
                resp.__exit__ = MagicMock(return_value=False)
                return resp
            # Second call: empty result → end
            resp = MagicMock()
            resp.read.return_value = _mock_find_people_response([])
            resp.__enter__ = lambda s: s
            resp.__exit__ = MagicMock(return_value=False)
            return resp

        with (
            patch("verificacion_correo.core.gal_scraper.urlopen", side_effect=mock_urlopen),
            patch("verificacion_correo.core.gal_scraper.time.sleep"),
        ):
            stats = scrape_gal(
                session_file=str(session_file),
                output_dir=output_dir,
                max_contacts=200,
                batch_size=100,
                request_delay=0,
            )

        # ProgressFile.save() does NOT persist the people list (only offset/count),
        # so resumed session only contains newly fetched contacts.
        assert stats["total"] == 2  # only new contacts from this run
        assert stats["offset_end"] == 152  # offset advanced from 150

        # Verify output files exist
        json_path = Path(output_dir) / OUTPUT_JSON
        csv_path = Path(output_dir) / OUTPUT_CSV
        assert json_path.exists()
        assert csv_path.exists()

        # Verify JSON content
        with open(json_path, encoding="utf-8") as f:
            data = json.load(f)
        assert len(data) == 2

    def test_scrape_gal_session_expired_saves_progress(self, tmp_path: Path) -> None:
        """scrape_gal guarda progreso cuando la sesión expira (HTTP 307)."""
        from urllib.error import HTTPError

        session_file = tmp_path / "state.json"
        _create_session_file(session_file)
        output_dir = str(tmp_path / "output")

        def mock_urlopen(req, timeout=60):
            raise HTTPError(req.full_url, 307, "Moved Temporarily", {}, None)

        with (
            patch("verificacion_correo.core.gal_scraper.urlopen", side_effect=mock_urlopen),
            patch("verificacion_correo.core.gal_scraper.time.sleep"),
        ):
            stats = scrape_gal(
                session_file=str(session_file),
                output_dir=output_dir,
                batch_size=100,
                request_delay=0,
            )

        assert stats["expired"] is True
        assert stats["total"] == 0

        # Progress file should exist so a resume is possible
        assert Path(output_dir, PROGRESS_FILENAME).exists()

    def test_scrape_gal_stop_flag(self, tmp_path: Path) -> None:
        """scrape_gal respeta stop_flag y guarda archivos."""
        session_file = tmp_path / "state.json"
        _create_session_file(session_file)
        output_dir = str(tmp_path / "output")

        def mock_urlopen(req, timeout=60):
            resp = MagicMock()
            resp.read.return_value = _mock_find_people_response([SAMPLE_PERSONA])
            resp.__enter__ = lambda s: s
            resp.__exit__ = MagicMock(return_value=False)
            return resp

        stop_flag = {"stop": True}

        with (
            patch("verificacion_correo.core.gal_scraper.urlopen", side_effect=mock_urlopen),
            patch("verificacion_correo.core.gal_scraper.time.sleep"),
        ):
            stats = scrape_gal(
                session_file=str(session_file),
                output_dir=output_dir,
                batch_size=100,
                request_delay=0,
                stop_flag=stop_flag,
            )

        assert stats["stopped"] is True

    def test_scrape_gal_empty_result_terminates(self, tmp_path: Path) -> None:
        """scrape_gal termina cuando la API retorna lista vacía."""
        session_file = tmp_path / "state.json"
        _create_session_file(session_file)
        output_dir = str(tmp_path / "output")

        def mock_urlopen(req, timeout=60):
            resp = MagicMock()
            resp.read.return_value = _mock_find_people_response([])
            resp.__enter__ = lambda s: s
            resp.__exit__ = MagicMock(return_value=False)
            return resp

        with (
            patch("verificacion_correo.core.gal_scraper.urlopen", side_effect=mock_urlopen),
            patch("verificacion_correo.core.gal_scraper.time.sleep"),
        ):
            stats = scrape_gal(
                session_file=str(session_file),
                output_dir=output_dir,
                batch_size=100,
                request_delay=0,
            )

        assert stats["total"] == 0
        assert Path(output_dir, OUTPUT_JSON).exists()


# ===========================================================================
# TestExcelReadWriteFlow
# ===========================================================================


class TestExcelReadWriteFlow:
    """Flujos de integración para ExcelReader y ExcelWriter."""

    def test_read_write_cycle(self, tmp_path: Path) -> None:
        """Excel read/write cycle: crear → leer → escribir → releer."""
        excel_path = str(tmp_path / "test.xlsx")

        # 1. Create file with headers via writer
        writer = ExcelWriter(excel_path)
        writer.ensure_file_structure()

        # 2. Read pending — file exists with headers only (no data rows)
        reader = ExcelReader(excel_path)
        summary = reader.read_pending_emails()
        assert summary.total_emails == 0
        assert summary.pending_count == 0

        # 3. Manually add rows with openpyxl
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        emails = ["user1@test.org", "user2@test.org", "user3@test.org"]
        for i, email in enumerate(emails, start=2):
            ws.cell(row=i, column=1, value=email)
        wb.save(excel_path)
        wb.close()

        # 4. Read pending again — should find all 3
        summary = reader.read_pending_emails()
        assert summary.total_emails == 3
        assert summary.pending_count == 3
        assert summary.processed_count == 0

        # 5. Write results for first two emails
        records = [
            EmailRecord(
                email="user1@test.org",
                row=2,
                status=ProcessingStatus.SUCCESS,
                data={"name": "User One", "phone": "111"},
            ),
            EmailRecord(
                email="user2@test.org",
                row=3,
                status=ProcessingStatus.NOT_FOUND,
            ),
        ]
        writer.write_batch_results(records)

        # 6. Read pending again — only 1 pending
        summary = reader.read_pending_emails()
        assert summary.total_emails == 3
        assert summary.pending_count == 1
        assert summary.processed_count == 2

        # 7. Verify written data
        records_read = reader.read_all_records()
        by_email = {r.email: r for r in records_read}

        assert by_email["user1@test.org"].status == ProcessingStatus.SUCCESS
        assert by_email["user1@test.org"].data["name"] == "User One"

        assert by_email["user2@test.org"].status == ProcessingStatus.ERROR
        assert by_email["user3@test.org"].status == ProcessingStatus.PENDING

    def test_read_pending_batching(self, tmp_path: Path) -> None:
        """read_pending_emails agrupa correctamente en batches."""
        excel_path = str(tmp_path / "batch_test.xlsx")

        writer = ExcelWriter(excel_path)
        writer.ensure_file_structure()

        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        for i in range(1, 8):
            ws.cell(row=i + 1, column=1, value=f"email{i}@test.org")
        wb.save(excel_path)
        wb.close()

        reader = ExcelReader(excel_path)
        summary = reader.read_pending_emails(batch_size=3)

        assert summary.total_emails == 7
        assert summary.pending_count == 7
        assert len(summary.batches) == 3  # 3 + 3 + 1
        assert len(summary.batches[0]) == 3
        assert len(summary.batches[2]) == 1

    def test_write_result_single_record(self, tmp_path: Path) -> None:
        """write_result escribe correctamente un registro individual."""
        excel_path = str(tmp_path / "single.xlsx")

        writer = ExcelWriter(excel_path)
        writer.ensure_file_structure()

        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=2, column=1, value="test@madrid.org")
        wb.save(excel_path)
        wb.close()

        record = EmailRecord(
            email="test@madrid.org",
            row=2,
            status=ProcessingStatus.SUCCESS,
            data={
                "name": "Test User",
                "email": "test.personal@gmail.com",
                "phone": "600123456",
                "sip": "sip:test@madrid.org",
                "department": "Engineering",
            },
        )
        writer.write_result(record)

        reader = ExcelReader(excel_path)
        records = reader.read_all_records()
        assert len(records) == 1
        assert records[0].status == ProcessingStatus.SUCCESS
        assert records[0].data["name"] == "Test User"
        assert records[0].data["sip"] == "sip:test@madrid.org"

    def test_read_all_emails(self, tmp_path: Path) -> None:
        """read_all_emails retorna todos los emails válidos."""
        excel_path = str(tmp_path / "emails.xlsx")

        writer = ExcelWriter(excel_path)
        writer.ensure_file_structure()

        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        ws = wb.active
        ws.cell(row=2, column=1, value="a@test.org")
        ws.cell(row=3, column=1, value="b@test.org")
        ws.cell(row=4, column=1, value="c@test.org")
        wb.save(excel_path)
        wb.close()

        reader = ExcelReader(excel_path)
        emails = reader.read_all_emails()
        assert emails == ["a@test.org", "b@test.org", "c@test.org"]
