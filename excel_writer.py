"""
Módulo para escribir resultados de scraping al archivo Excel.
"""

import os
from openpyxl import load_workbook, Workbook


# Definición de columnas
COLUMNAS = {
    'EMAIL': 1,        # A
    'STATUS': 2,       # B
    'NOMBRE': 3,       # C
    'EMAIL_PERSONAL': 4,  # D
    'TELEFONO': 5,     # E
    'SIP': 6,          # F
    'DIRECCION': 7,    # G
    'DEPARTAMENTO': 8, # H
    'COMPANIA': 9,     # I
    'OFICINA': 10      # J
}

HEADERS = [
    'Correo',
    'Status',
    'Nombre',
    'Email Personal',
    'Teléfono',
    'SIP',
    'Dirección',
    'Departamento',
    'Compañía',
    'Oficina'
]


def crear_headers_si_no_existen(archivo_path):
    """
    Verifica que el archivo Excel tenga los headers correctos.
    Si no existen o están incompletos, los crea/completa.

    Args:
        archivo_path: Ruta al archivo Excel
    """
    # Si el archivo no existe, crear uno nuevo con headers
    if not os.path.exists(archivo_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"

        # Escribir headers en la fila 1
        for col_idx, header in enumerate(HEADERS, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Crear el directorio si no existe
        os.makedirs(os.path.dirname(archivo_path), exist_ok=True)
        wb.save(archivo_path)
        return

    # Si existe, verificar que tenga headers completos
    wb = load_workbook(archivo_path)
    ws = wb.active

    # Verificar y completar headers si es necesario
    headers_completos = True
    for col_idx, header in enumerate(HEADERS, start=1):
        celda = ws.cell(row=1, column=col_idx)
        if celda.value != header:
            celda.value = header
            headers_completos = False

    if not headers_completos:
        wb.save(archivo_path)

    wb.close()


def escribir_resultado(archivo_path, fila, status, datos=None):
    """
    Escribe el resultado de un scraping a una fila específica del Excel.

    Args:
        archivo_path: Ruta al archivo Excel
        fila: Número de fila donde escribir (base 1, fila 2 es la primera después del header)
        status: Estado del procesamiento ('OK', 'NO EXISTE', 'ERROR')
        datos: Diccionario con los datos extraídos (opcional)
            Formato: {
                'name': str,
                'email': str,
                'phone': str,
                'sip': str,
                'address': str,
                'department': str,
                'company': str,
                'office_location': str
            }
    """
    # Asegurar que el archivo y headers existan
    crear_headers_si_no_existen(archivo_path)

    # Abrir el archivo
    wb = load_workbook(archivo_path)
    ws = wb.active

    # Escribir el status
    ws.cell(row=fila, column=COLUMNAS['STATUS'], value=status)

    # Si hay datos, escribirlos
    if datos and status == 'OK':
        ws.cell(row=fila, column=COLUMNAS['NOMBRE'], value=datos.get('name'))
        ws.cell(row=fila, column=COLUMNAS['EMAIL_PERSONAL'], value=datos.get('email'))
        ws.cell(row=fila, column=COLUMNAS['TELEFONO'], value=datos.get('phone'))
        ws.cell(row=fila, column=COLUMNAS['SIP'], value=datos.get('sip'))
        ws.cell(row=fila, column=COLUMNAS['DIRECCION'], value=datos.get('address'))
        ws.cell(row=fila, column=COLUMNAS['DEPARTAMENTO'], value=datos.get('department'))
        ws.cell(row=fila, column=COLUMNAS['COMPANIA'], value=datos.get('company'))
        ws.cell(row=fila, column=COLUMNAS['OFICINA'], value=datos.get('office_location'))
    elif status in ['NO EXISTE', 'ERROR']:
        # Limpiar las columnas de datos si el status es error o no existe
        for col in range(COLUMNAS['NOMBRE'], COLUMNAS['OFICINA'] + 1):
            ws.cell(row=fila, column=col, value=None)

    # Guardar el archivo
    wb.save(archivo_path)
    wb.close()


def leer_status(archivo_path, fila):
    """
    Lee el status de una fila específica.

    Args:
        archivo_path: Ruta al archivo Excel
        fila: Número de fila (base 1)

    Returns:
        El valor de la columna Status, o None si está vacía
    """
    if not os.path.exists(archivo_path):
        return None

    wb = load_workbook(archivo_path, read_only=True, data_only=True)
    ws = wb.active

    status = ws.cell(row=fila, column=COLUMNAS['STATUS']).value

    wb.close()

    return status
