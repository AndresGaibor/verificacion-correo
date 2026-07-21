"""
Main GUI module for verificacion-correo.

This module provides a modern Tkinter-based graphical user interface for
email verification and contact extraction from OWA.
"""

import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
import threading
import queue
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, Dict, Any, Callable
import yaml

from verificacion_correo.core.platform import open_file, open_folder

from verificacion_correo.core.config import Config
from verificacion_correo.core.excel import ExcelReader
from verificacion_correo.utils.logging import setup_logging, get_logger
from verificacion_correo.gui.service import GUIService
from verificacion_correo.gui.wizard import ConfigWizard


logger = get_logger(__name__)


class ToolTip:
    """Tooltip helper for tkinter widgets."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind('<Enter>', self._show)
        widget.bind('<Leave>', self._hide)

    def _show(self, event=None):
        if self.tip_window or not self.text:
            return
        x = self.widget.winfo_rootx() + 20
        y = self.widget.winfo_rooty() + 20
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)
        tw.wm_geometry(f"+{x}+{y}")
        label = tk.Label(tw, text=self.text, background="#ffffe0", relief="solid", borderwidth=1, padx=4, pady=2)
        label.pack()

    def _hide(self, event=None):
        if self.tip_window:
            self.tip_window.destroy()
        self.tip_window = None


class VerificacionCorreosGUI:
    """Main GUI application class."""

    def __init__(self, root: tk.Tk):
        """Initialize GUI."""
        self.root = root
        self.root.title("Verificación de Correos OWA v2.0")
        self.root.geometry("1000x700")
        self.root.minsize(800, 600)

        # Load configuration
        try:
            self.config = Config()
        except Exception as e:
            messagebox.showerror("Configuration Error", f"Failed to load config: {e}")
            self.root.quit()
            return

        # Initialize service
        self.service = GUIService(self.config)

        # Setup logging for GUI - use DEBUG to see detailed asyncio/playwright logs
        setup_logging(level="DEBUG")
        self.log_messages = []
        
        self.scraper_output_dir = tk.StringVar(value=str(Path.cwd() / "data"))
        self.scraper_max_contacts = tk.IntVar(value=500)
        self.scraper_extracted_count = tk.IntVar(value=0)
        self.is_processing = False
        self.scraper_active = False
        self.scraper_log_messages = []

        # Session confirmation dialog state
        self._session_confirm_window = None
        self._confirm_status = None
        self._session_confirm_closing = False

        # Address list selection
        self.address_lists = []  # List of {"DisplayName": ..., "FolderId": {"Id": ...}}
        self.selected_address_list_id = tk.StringVar(value="fed75805-8ba2-4323-9f6d-80be7e3abc6a")

        # Company filter variables
        self.all_companies = []
        self.company_search_var = tk.StringVar(value="")

        # Create interface
        self._create_widgets()
        self._setup_status_check()
        self._setup_keyboard_shortcuts()
        self._setup_safe_close()

    def _create_widgets(self):
        """Create all GUI widgets."""
        # Create main container with padding
        main_container = ttk.Frame(self.root)
        main_container.pack(fill='both', expand=True, padx=10, pady=10)

        # Create menu bar
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)
        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="Ayuda", menu=help_menu)
        help_menu.add_command(label="Acerca de...", command=self._show_about)

        # Create notebook for tabs
        self.notebook = ttk.Notebook(main_container)
        self.notebook.pack(fill='both', expand=True)

        # Create tabs
        self._create_processing_tab()
        self._create_session_tab()
        self._create_config_tab()
        self._create_scraper_tab()

        # Create status bar
        self._create_status_bar(main_container)

        # Bind tab change
        self.notebook.bind('<<NotebookTabChanged>>', self._on_tab_changed)

    def _create_processing_tab(self):
        """Create processing tab."""
        self.processing_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.processing_frame, text="📧 Procesamiento")

        # Main container
        main_frame = ttk.Frame(self.processing_frame)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="Archivo de Excel", padding=10)
        file_frame.pack(fill='x', pady=(0, 10))

        self.excel_path_var = tk.StringVar(value=str(self.config.get_excel_file_path()))
        ttk.Entry(file_frame, textvariable=self.excel_path_var, width=60).pack(side='left', padx=(0, 5))
        ttk.Button(file_frame, text="Seleccionar", command=self._select_excel_file).pack(side='left')
        btn_refrescar = ttk.Button(file_frame, text="Refrescar", command=self._refresh_excel_info).pack(side='left', padx=(5, 0))

        # Summary section
        summary_frame = ttk.LabelFrame(main_frame, text="Resumen", padding=10)
        summary_frame.pack(fill='x', pady=(0, 10))

        self.summary_text = tk.StringVar(value="Cargando información...")
        ttk.Label(summary_frame, textvariable=self.summary_text, wraplength=600).pack()

        # Automation engine indicator
        engine_info = self._get_automation_engine_info()
        self.engine_label = ttk.Label(
            summary_frame,
            text=engine_info,
            foreground='gray'
        )
        self.engine_label.pack(pady=(5, 0))

        # Control buttons
        control_frame = ttk.Frame(main_frame)
        control_frame.pack(fill='x', pady=(0, 10))

        self.start_btn = ttk.Button(
            control_frame,
            text="🚀 Iniciar Procesamiento",
            command=self._start_processing,
            style='Accent.TButton'
        )
        self.start_btn.pack(side='left', padx=(0, 5))
        ToolTip(self.start_btn, "Inicia el procesamiento de correos pendientes (Ctrl+Enter)")

        self.stop_btn = ttk.Button(
            control_frame,
            text="⏹ Detener",
            command=self._stop_processing,
            state='disabled'
        )
        self.stop_btn.pack(side='left', padx=5)
        ToolTip(self.stop_btn, "Detiene el procesamiento en curso (Escape)")

        btn_ver = ttk.Button(
            control_frame,
            text="📊 Ver Resultados",
            command=self._open_excel_file
        )
        btn_ver.pack(side='left', padx=(5, 0))
        ToolTip(btn_ver, "Abre el archivo Excel con los resultados")

        self.api_btn = ttk.Button(
            control_frame,
            text="🔍 Buscar por API",
            command=self._start_api_search
        )
        self.api_btn.pack(side='left', padx=(5, 0))
        ToolTip(self.api_btn, "Busca contactos en el directorio OWA vía API REST")

        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Progreso", padding=10)
        progress_frame.pack(fill='x', pady=(0, 10))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(
            progress_frame,
            variable=self.progress_var,
            maximum=100,
            mode='determinate'
        )
        self.progress_bar.pack(fill='x', pady=(0, 5))

        self.progress_text = tk.StringVar(value="Listo para procesar")
        ttk.Label(progress_frame, textvariable=self.progress_text).pack()

        # Results section
        results_frame = ttk.LabelFrame(main_frame, text="Resultados", padding=10)
        results_frame.pack(fill='x', pady=(0, 10))

        columns = ('correo', 'status', 'nombre', 'email_personal', 'telefono')
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show='headings', height=5)
        self.results_tree.heading('correo', text='Correo')
        self.results_tree.heading('status', text='Estado')
        self.results_tree.heading('nombre', text='Nombre')
        self.results_tree.heading('email_personal', text='Email Personal')
        self.results_tree.heading('telefono', text='Teléfono')
        self.results_tree.column('correo', width=200)
        self.results_tree.column('status', width=80)
        self.results_tree.column('nombre', width=150)
        self.results_tree.column('email_personal', width=180)
        self.results_tree.column('telefono', width=120)
        self.results_tree.pack(fill='x')

        results_scroll = ttk.Scrollbar(results_frame, orient='vertical', command=self.results_tree.yview)
        self.results_tree.configure(yscrollcommand=results_scroll.set)
        results_scroll.pack(side='right', fill='y')

        # Log section
        log_frame = ttk.LabelFrame(main_frame, text="Registro de Eventos", padding=10)
        log_frame.pack(fill='both', expand=True)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            height=15,
            wrap=tk.WORD,
            font=('Consolas', 9)
        )
        self.log_text.pack(fill='both', expand=True)

        # Log control buttons
        log_control_frame = ttk.Frame(log_frame)
        log_control_frame.pack(fill='x', pady=(5, 0))

        ttk.Button(log_control_frame, text="Limpiar", command=self._clear_log).pack(side='left')
        ttk.Button(log_control_frame, text="Guardar Log", command=self._save_log).pack(side='left', padx=(5, 0))

    def _create_session_tab(self):
        """Create session management tab."""
        self.session_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.session_frame, text="🔐 Sesión del Navegador")

        main_frame = ttk.Frame(self.session_frame)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Session status
        status_frame = ttk.LabelFrame(main_frame, text="Estado de la Sesión", padding=10)
        status_frame.pack(fill='x', pady=(0, 10))

        self.session_status_text = tk.StringVar(value="Verificando...")
        ttk.Label(status_frame, textvariable=self.session_status_text, wraplength=600).pack()

        # Session actions - dynamic button
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill='x', pady=(0, 10))

        self.session_action_btn = ttk.Button(
            action_frame,
            text="🔐 Iniciar Sesión",
            command=self._handle_session_action
        )
        self.session_action_btn.pack(side='left', padx=(0, 5))

        btn_eliminar = ttk.Button(
            action_frame,
            text="🗑️ Eliminar Sesión",
            command=self._delete_session
        )
        btn_eliminar.pack(side='left', padx=(5, 0))

        # Session info
        info_frame = ttk.LabelFrame(main_frame, text="Información de la Sesión", padding=10)
        info_frame.pack(fill='both', expand=True)

        self.session_info_text = scrolledtext.ScrolledText(
            info_frame,
            height=15,
            wrap=tk.WORD,
            font=('Consolas', 9),
            state='disabled'
        )
        self.session_info_text.pack(fill='both', expand=True)

    def _create_config_tab(self):
        """Create configuration tab."""
        self.config_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.config_frame, text="⚙️ Configuración")

        main_frame = ttk.Frame(self.config_frame)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Configuration editor
        editor_frame = ttk.LabelFrame(main_frame, text="Editor de Configuración", padding=10)
        editor_frame.pack(fill='both', expand=True, pady=(0, 10))

        # Create scrollable configuration editor
        self._create_config_editor(editor_frame)

        # Quick actions
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill='x', pady=(0, 10))

        btn_guardar = ttk.Button(
            action_frame,
            text="💾 Guardar Configuración",
            command=self._save_config
        )
        btn_guardar.pack(side='left', padx=(0, 5))
        ToolTip(btn_guardar, "Guarda la configuración actual (Ctrl+S)")

        btn_reargar = ttk.Button(
            action_frame,
            text="🔄 Recargar Configuración",
            command=self._reload_config
        )
        btn_reargar.pack(side='left', padx=(5, 0))
        ToolTip(btn_reargar, "Recarga la configuración desde el archivo")

        btn_carpeta = ttk.Button(
            action_frame,
            text="📁 Abrir Carpeta de Datos",
            command=self._open_data_folder
        )
        btn_carpeta.pack(side='left', padx=(5, 0))
        ToolTip(btn_carpeta, "Abre la carpeta donde se guardan los datos")

        btn_asistente = ttk.Button(
            action_frame,
            text="🔧 Asistente de Configuración",
            command=self._run_config_wizard
        )
        btn_asistente.pack(side='left', padx=(5, 0))
        ToolTip(btn_asistente, "Abre el asistente de configuración inicial")

    def _create_scraper_tab(self):
        """Create scraper tab for extracting Outlook contacts."""
        self.scraper_frame = ttk.Frame(self.notebook)
        self.notebook.add(self.scraper_frame, text="🔍 Scraper de Contactos")

        main_frame = ttk.Frame(self.scraper_frame)
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)

        # Configuration section
        config_frame = ttk.LabelFrame(main_frame, text="Configuración de Scraping", padding=10)
        config_frame.pack(fill='x', pady=(0, 10))

        # Output directory
        dir_label_frame = ttk.Frame(config_frame)
        dir_label_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(dir_label_frame, text="Directorio de salida:").pack(anchor='w')

        dir_input_frame = ttk.Frame(config_frame)
        dir_input_frame.pack(fill='x', pady=(0, 10))

        ttk.Entry(
            dir_input_frame,
            textvariable=self.scraper_output_dir,
            state='readonly',
            width=60
        ).pack(side='left', fill='x', expand=True, padx=(0, 5))

        ttk.Button(
            dir_input_frame,
            text="📁 Seleccionar",
            command=self._select_scraper_directory
        ).pack(side='left')

        # Contact quantity
        quantity_label_frame = ttk.Frame(config_frame)
        quantity_label_frame.pack(fill='x', pady=(0, 5))
        ttk.Label(quantity_label_frame, text="Cantidad de contactos a extraer:").pack(anchor='w')

        quantity_frame = ttk.Frame(config_frame)
        quantity_frame.pack(fill='x')

        vcmd_scraper = (self.root.register(self._validate_numeric), '%P')
        ttk.Spinbox(
            quantity_frame,
            from_=1,
            to=10000,
            textvariable=self.scraper_max_contacts,
            width=20,
            validate='key',
            validatecommand=vcmd_scraper
        ).pack(anchor='w')

        # Address list selection
        self.addr_frame = ttk.LabelFrame(config_frame, text="Lista de Direcciones (Address List)", padding=5)
        self.addr_frame.pack(fill='x', pady=(5, 0))

        addr_controls = ttk.Frame(self.addr_frame)
        addr_controls.pack(fill='x')

        self.address_list_combo = ttk.Combobox(
            addr_controls,
            textvariable=self.selected_address_list_id,
            state='readonly',
            width=60,
        )
        self.address_list_combo.pack(side='left', fill='x', expand=True, padx=(0, 5))

        self.addr_list_scan_btn = ttk.Button(
            addr_controls,
            text="🔄 Cargar listas",
            command=self._scan_address_lists,
        )
        self.addr_list_scan_btn.pack(side='left')

        self.addr_list_status_label = ttk.Label(
            self.addr_frame,
            text="Use 'Cargar listas' para descubrir las listas de direcciones disponibles en OWA",
            foreground='gray',
            font=('TkDefaultFont', 9, 'italic'),
        )
        self.addr_list_status_label.pack(anchor='w', pady=(3, 0))

        # Company filter section
        self.filter_frame = ttk.LabelFrame(main_frame, text="Filtro de Compañías", padding=10)
        self.filter_frame.pack(fill='x', pady=(0, 10))
        self.filter_frame.pack_forget()  # Hidden: feature not working

        # Address list also hidden
        self.addr_frame.pack_forget()  # Hidden: feature not working

        # Enable/disable filter
        self.company_filter_enabled = tk.BooleanVar(
            value=self.config.company_filter.enabled if hasattr(self.config, 'company_filter') else False
        )
        ttk.Checkbutton(
            self.filter_frame,
            text="Filtrar por compañía (solo guardar contactos de las compañías seleccionadas)",
            variable=self.company_filter_enabled,
            command=self._on_company_filter_toggle
        ).pack(anchor='w', pady=(0, 5))

        # Company filter controls
        filter_controls = ttk.Frame(self.filter_frame)
        filter_controls.pack(fill='x', pady=(0, 5))

        self.company_scan_btn = ttk.Button(
            filter_controls,
            text="🔄 Cargar compañías desde API",
            command=self._scan_companies
        )
        self.company_scan_btn.pack(side='left', padx=(0, 5))

        # Company search box
        search_entry = ttk.Entry(
            filter_controls,
            textvariable=self.company_search_var,
            width=25,
        )
        search_entry.pack(side='left', padx=(0, 3))
        search_entry.insert(0, "🔍 Buscar compañía...")
        search_entry.config(foreground='gray')
        search_entry.bind('<FocusIn>', lambda e: self._on_search_focus_in(search_entry))
        search_entry.bind('<FocusOut>', lambda e: self._on_search_focus_out(search_entry))
        search_entry.bind('<KeyRelease>', lambda e: self._filter_company_list())
        self.company_search_entry = search_entry

        # Manual company entry
        ttk.Label(filter_controls, text="+").pack(side='left', padx=(5, 2))
        self.company_manual_entry = ttk.Entry(filter_controls, width=22)
        self.company_manual_entry.pack(side='left', padx=(0, 3))
        self.company_manual_entry.insert(0, "Escribir compañía...")
        self.company_manual_entry.config(foreground='gray')
        self.company_manual_entry.bind('<FocusIn>', lambda e: self._on_manual_company_focus_in())
        self.company_manual_entry.bind('<FocusOut>', lambda e: self._on_manual_company_focus_out())
        self.company_manual_entry.bind('<Return>', lambda e: self._add_manual_company())

        ttk.Button(
            filter_controls,
            text="+",
            width=3,
            command=self._add_manual_company
        ).pack(side='left', padx=(0, 5))

        ttk.Button(
            filter_controls,
            text="Seleccionar todas",
            command=self._select_all_companies
        ).pack(side='left', padx=(0, 5))

        ttk.Button(
            filter_controls,
            text="Deseleccionar todas",
            command=self._deselect_all_companies
        ).pack(side='left')

        # Company list with scrollbar
        list_frame = ttk.Frame(self.filter_frame)
        list_frame.pack(fill='x')

        self.company_list_canvas = tk.Canvas(list_frame, height=120, highlightthickness=0)
        company_scrollbar = ttk.Scrollbar(list_frame, orient='vertical', command=self.company_list_canvas.yview)
        self.company_list_inner = ttk.Frame(self.company_list_canvas)

        self.company_list_inner.bind(
            '<Configure>',
            lambda e: self.company_list_canvas.configure(scrollregion=self.company_list_canvas.bbox("all"))
        )
        self.company_list_canvas.create_window((0, 0), window=self.company_list_inner, anchor='nw')
        self.company_list_canvas.configure(yscrollcommand=company_scrollbar.set)

        self.company_list_canvas.pack(side='left', fill='both', expand=True)
        company_scrollbar.pack(side='right', fill='y')

        # Store company checkboxes
        self.company_checkboxes = {}
        self.company_vars = {}

        # Company status label
        self.company_status_label = ttk.Label(
            self.filter_frame,
            text="No hay compañías cargadas",
            foreground='gray',
            font=('TkDefaultFont', 9, 'italic')
        )
        self.company_status_label.pack(anchor='w', pady=(5, 0))

        # Control section
        control_frame = ttk.LabelFrame(main_frame, text="Control de Extracción", padding=10)
        control_frame.pack(fill='x', pady=(0, 10))

        button_frame = ttk.Frame(control_frame)
        button_frame.pack(fill='x')

        self.scraper_start_btn = ttk.Button(
            button_frame,
            text="▶️ Extraer Directorio",
            command=self._start_scraper,
            style='Accent.TButton'
        )
        self.scraper_start_btn.pack(side='left', padx=(0, 5), fill='x', expand=True)
        ToolTip(self.scraper_start_btn, "Extrae directorio GAL completo a Excel (Ctrl+Shift+Enter)")

        self.scraper_stop_btn = ttk.Button(
            button_frame,
            text="⏹️ Detener",
            command=self._stop_scraper,
            state='disabled'
        )
        self.scraper_stop_btn.pack(side='left', fill='x', expand=True)
        ToolTip(self.scraper_stop_btn, "Detiene la extracción en curso (Escape)")

        self.enrich_btn = ttk.Button(
            button_frame,
            text="🔄 Completar información",
            command=self._start_enrichment
        )
        self.enrich_btn.pack(side='left', padx=(5, 0))
        ToolTip(self.enrich_btn, "Busca teléfono, departamento y dirección de contactos usando GetPersona API")

        btn_abrir_resultados = ttk.Button(
            button_frame,
            text="📂 Abrir",
            command=self._open_scraper_output
        )
        btn_abrir_resultados.pack(side='left', padx=(5, 0))
        ToolTip(btn_abrir_resultados, "Abre carpeta de resultados")

        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Progreso de Extracción", padding=10)
        progress_frame.pack(fill='x', pady=(0, 10))

        # Counter
        counter_frame = ttk.Frame(progress_frame)
        counter_frame.pack(fill='x', pady=(0, 5))

        ttk.Label(counter_frame, text="Contactos extraídos:").pack(side='left')

        self.scraper_counter_label = ttk.Label(
            counter_frame,
            textvariable=self.scraper_extracted_count,
            font=('TkDefaultFont', 14, 'bold'),
            foreground='#27ae60'
        )
        self.scraper_counter_label.pack(side='left', padx=5)

        self.scraper_total_label = ttk.Label(
            counter_frame,
            text="/ 0",
            font=('TkDefaultFont', 12),
            foreground='gray'
        )
        self.scraper_total_label.pack(side='left')

        # Progress bar
        self.scraper_progress_bar = ttk.Progressbar(
            progress_frame,
            mode='determinate',
            length=300
        )
        self.scraper_progress_bar.pack(fill='x', pady=(0, 5))

        # Status
        self.scraper_status_label = ttk.Label(
            progress_frame,
            text="⚪ Esperando...",
            font=('TkDefaultFont', 10, 'italic'),
            foreground='gray'
        )
        self.scraper_status_label.pack(anchor='w')

        # Log section
        log_frame = ttk.LabelFrame(main_frame, text="Log de Scraping", padding=10)
        log_frame.pack(fill='both', expand=True)

        self.scraper_log_text = scrolledtext.ScrolledText(
            log_frame,
            height=12,
            wrap=tk.WORD,
            font=('Consolas', 9)
        )
        self.scraper_log_text.pack(fill='both', expand=True, pady=(0, 5))

        # Log control buttons
        log_control_frame = ttk.Frame(log_frame)
        log_control_frame.pack(fill='x')

        ttk.Button(
            log_control_frame,
            text="Limpiar",
            command=self._clear_scraper_log
        ).pack(side='left', padx=(0, 5))

        ttk.Button(
            log_control_frame,
            text="Guardar Log",
            command=self._save_scraper_log
        ).pack(side='left')

        # Initial log message
        self._add_scraper_log("✅ Interfaz de scraper iniciada correctamente")
        self._add_scraper_log(f"📁 Directorio de salida: {self.scraper_output_dir.get()}")
        self._update_scraper_total_label()

        # Load cached companies if available
        self._load_cached_companies()

    def _select_scraper_directory(self):
        """Select output directory for scraper."""
        directory = filedialog.askdirectory(
            title="Seleccionar directorio de salida",
            initialdir=self.scraper_output_dir.get()
        )
        if directory:
            self.scraper_output_dir.set(directory)
            self._add_scraper_log(f"📁 Directorio cambi ado a: {directory}")

    def _update_scraper_total_label(self):
        """Update the total label with current max contacts."""
        total = self.scraper_max_contacts.get()
        self.scraper_total_label.config(text=f"/ {total}")

    def _add_scraper_log(self, message):
        """Add message to scraper log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.scraper_log_messages.append(log_entry)

        self.scraper_log_text.insert(tk.END, log_entry)
        self.scraper_log_text.see(tk.END)

    def _clear_scraper_log(self):
        """Clear scraper log messages."""
        self.scraper_log_text.delete('1.0', tk.END)
        self.scraper_log_messages.clear()

    def _save_scraper_log(self):
        """Save scraper log messages to file."""
        file_path = filedialog.asksaveasfilename(
            title="Guardar Log",
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.writelines(self.scraper_log_messages)
                messagebox.showinfo("Log Guardado", f"Log guardado en:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar log: {e}")

    def _update_scraper_progress(self, count):
        """Update scraper progress counter and bar."""
        self.scraper_extracted_count.set(count)
        total = self.scraper_max_contacts.get()

        # Update progress bar
        if total > 0:
            progress = (count / total) * 100
            self.scraper_progress_bar['value'] = progress

    def _update_scraper_status(self, message, color='gray'):
        """Update scraper status label."""
        self.scraper_status_label.config(text=message, foreground=color)

    # --- Address list methods ---

    def _scan_address_lists(self):
        """Start address list scan from OWA API."""
        session_file = self.config.get_session_file_path()
        if not Path(session_file).exists():
            messagebox.showwarning(
                "Sesión Requerida",
                "No hay sesión guardada. Configura una sesión primero."
            )
            return

        self.addr_list_scan_btn.config(state='disabled')
        self.addr_list_status_label.config(text="🔄 Escaneando listas de direcciones...", foreground='#3498db')
        self._add_scraper_log("🔄 Iniciando escaneo de listas de direcciones vía GetPeopleFilters...")

        try:
            self.service.start_address_list_scan()
        except Exception as e:
            self._add_scraper_log(f"❌ Error al iniciar escaneo de listas: {e}")
            self.addr_list_scan_btn.config(state='normal')
            self.addr_list_status_label.config(text="Error al escanear", foreground='#e74c3c')

    def _handle_address_list_scan_complete(self, result):
        """Handle address list scan completion."""
        self.root.after(0, lambda: self._on_address_list_scan_complete(result))

    def _on_address_list_scan_complete(self, result):
        """Process address list scan results in GUI thread."""
        address_lists = result.get('address_lists', [])
        count = result.get('count', 0)

        self.addr_list_scan_btn.config(state='normal')
        self.address_lists = address_lists

        if address_lists:
            # Build display names for combobox
            display_names = [f"{al['DisplayName']}" for al in address_lists]
            self.address_list_combo['values'] = display_names

            # Try to select the default GAL
            default_idx = 0
            for i, al in enumerate(address_lists):
                if 'default' in al['DisplayName'].lower() and 'global' in al['DisplayName'].lower():
                    default_idx = i
                    break

            self.address_list_combo.current(default_idx)
            self._update_selected_address_list_id()

            self.addr_list_status_label.config(
                text=f"✅ {count} listas encontradas — seleccionada: {address_lists[default_idx]['DisplayName']}",
                foreground='#27ae60',
            )
            self._add_scraper_log(f"✅ Escaneo completado: {count} listas de direcciones encontradas")
            for al in address_lists:
                self._add_scraper_log(f"   📋 {al['DisplayName']} → {al['FolderId']['Id']}")
        else:
            self.addr_list_status_label.config(text="⚠️ No se encontraron listas", foreground='#e67e22')
            self._add_scraper_log("⚠️ No se encontraron listas de direcciones")

    def _handle_address_list_scan_error(self, error_msg):
        """Handle address list scan error."""
        self.root.after(0, lambda: self._on_address_list_scan_error(error_msg))

    def _on_address_list_scan_error(self, error_msg):
        """Process address list scan error in GUI thread."""
        self.addr_list_scan_btn.config(state='normal')
        self.addr_list_status_label.config(text="Error al escanear", foreground='#e74c3c')
        self._add_scraper_log(f"❌ Error en escaneo de listas: {error_msg}")
        messagebox.showerror("Error de Escaneo", f"Error al escanear listas de direcciones:\n{error_msg}")

    def _update_selected_address_list_id(self):
        """Update the selected AddressListId based on combobox selection."""
        selection = self.address_list_combo.current()
        if selection >= 0 and selection < len(self.address_lists):
            list_id = self.address_lists[selection]['FolderId']['Id']
            self.selected_address_list_id.set(list_id)

    def _on_search_focus_in(self, entry):
        """Handle search entry focus in."""
        if entry.get() == "🔍 Buscar compañía...":
            entry.delete(0, tk.END)
            entry.config(foreground='black')

    def _on_search_focus_out(self, entry):
        """Handle search entry focus out."""
        if entry.get().strip() == "":
            entry.insert(0, "🔍 Buscar compañía...")
            entry.config(foreground='gray')
            self._filter_company_list()

    def _on_manual_company_focus_in(self):
        entry = self.company_manual_entry
        if entry.get() == "Escribir compañía...":
            entry.delete(0, tk.END)
            entry.config(foreground='black')

    def _on_manual_company_focus_out(self):
        entry = self.company_manual_entry
        if entry.get().strip() == "":
            entry.insert(0, "Escribir compañía...")
            entry.config(foreground='gray')

    def _filter_company_list(self):
        """Filter company list based on search text."""
        search_text = self.company_search_var.get().strip().lower()
        if search_text == "" or search_text == "🔍 buscar compañía...":
            search_text = ""

        for company, var in self.company_vars.items():
            if not search_text or search_text in company.lower():
                self.company_checkboxes[company].pack_forget()
                self.company_checkboxes[company].pack(anchor='w', padx=5)
            else:
                self.company_checkboxes[company].pack_forget()

        self.company_list_canvas.update_idletasks()
        self.company_list_canvas.configure(
            scrollregion=self.company_list_canvas.bbox("all")
        )

    def _add_manual_company(self):
        """Add a manually entered company to the filter."""
        entry = self.company_manual_entry
        text = entry.get().strip()

        if not text or text == "Escribir compañía...":
            return

        # Split by comma or newline
        new_companies = [c.strip() for c in text.replace('\n', ',').split(',') if c.strip()]

        for company_name in new_companies:
            if company_name not in self.company_vars:
                # Create checkbox for this new company
                var = tk.BooleanVar(value=True)
                self.company_vars[company_name] = var
                cb = ttk.Checkbutton(
                    self.company_list_inner,
                    text=f"✏️ {company_name} (manual)",
                    variable=var,
                    command=self._save_company_selection
                )
                cb.pack(anchor='w', padx=5)
                self.company_checkboxes[company_name] = cb
                self.all_companies.append(company_name)
                self._add_scraper_log(f"   ➕ Compañía manual agregada: {company_name}")
            else:
                # Already exists, just check it
                self.company_vars[company_name].set(True)

        # Clear entry
        entry.delete(0, tk.END)
        entry.insert(0, "Escribir compañía...")
        entry.config(foreground='gray')

        self._save_company_selection()
        self._update_company_status()

    # --- Company filter methods ---

    def _load_cached_companies(self):
        """Load cached company list from previous scan."""
        try:
            companies = self.service.get_cached_companies()
            if companies:
                self._populate_company_list(companies)
                self._add_scraper_log(f"📋 {len(companies)} compañías cargadas desde caché")
        except Exception as e:
            logger.debug(f"No se pudo cargar caché de compañías: {e}")

    def _scan_companies(self):
        """Start company list scan from API."""
        session_file = self.config.get_session_file_path()
        if not Path(session_file).exists():
            messagebox.showwarning(
                "Sesión Requerida",
                "No hay sesión guardada. Configura una sesión primero."
            )
            return

        self.company_scan_btn.config(state='disabled')
        self.company_status_label.config(text="🔄 Escaneando compañías...", foreground='#3498db')
        self._add_scraper_log("🔄 Iniciando escaneo de compañías vía API...")

        try:
            self.service.start_company_scan(address_list_id=self.selected_address_list_id.get())
        except Exception as e:
            self._add_scraper_log(f"❌ Error al iniciar escaneo: {e}")
            self.company_scan_btn.config(state='normal')
            self.company_status_label.config(text="Error al escanear", foreground='#e74c3c')

    def _handle_company_scan_complete(self, result):
        """Handle company scan completion."""
        self.root.after(0, lambda: self._on_company_scan_complete(result))

    def _on_company_scan_complete(self, result):
        """Process company scan results in GUI thread."""
        companies = result.get('companies', [])
        count = result.get('count', 0)

        self.company_scan_btn.config(state='normal')
        self.company_status_label.config(
            text=f"✅ {count} compañías encontradas",
            foreground='#27ae60'
        )
        self._add_scraper_log(f"✅ Escaneo completado: {count} compañías encontradas")

        if companies:
            self._populate_company_list(companies)
            # Save to cache
            output_dir = Path(self.config.get_excel_file_path()).parent / "gal"
            output_dir.mkdir(parents=True, exist_ok=True)
            from verificacion_correo.core.gal_scraper import save_companies_cache
            save_companies_cache(companies, output_dir)

    def _handle_company_scan_error(self, error_msg):
        """Handle company scan error."""
        self.root.after(0, lambda: self._on_company_scan_error(error_msg))

    def _on_company_scan_error(self, error_msg):
        """Process company scan error in GUI thread."""
        self.company_scan_btn.config(state='normal')
        self.company_status_label.config(text="Error al escanear", foreground='#e74c3c')
        self._add_scraper_log(f"❌ Error en escaneo: {error_msg}")
        messagebox.showerror("Error de Escaneo", f"Error al escanear compañías:\n{error_msg}")

    def _populate_company_list(self, companies):
        """Populate the company checkbox list."""
        # Clear existing
        for widget in self.company_list_inner.winfo_children():
            widget.destroy()
        self.company_checkboxes.clear()
        self.company_vars.clear()

        # Load saved selections
        saved_companies = set()
        if hasattr(self.config, 'company_filter') and self.config.company_filter.companies:
            saved_companies = set(self.config.company_filter.companies)

        for company in companies:
            var = tk.BooleanVar(value=company in saved_companies if saved_companies else True)
            self.company_vars[company] = var
            cb = ttk.Checkbutton(
                self.company_list_inner,
                text=company,
                variable=var,
                command=self._save_company_selection
            )
            cb.pack(anchor='w', padx=5)
            self.company_checkboxes[company] = cb

        # Update status
        selected = sum(1 for v in self.company_vars.values() if v.get())
        self.company_status_label.config(
            text=f"✅ {len(companies)} compañías — {selected} seleccionadas",
            foreground='#27ae60'
        )

    def _select_all_companies(self):
        """Select all companies in the list."""
        for var in self.company_vars.values():
            var.set(True)
        self._save_company_selection()
        self._update_company_status()

    def _deselect_all_companies(self):
        """Deselect all companies in the list."""
        for var in self.company_vars.values():
            var.set(False)
        self._save_company_selection()
        self._update_company_status()

    def _save_company_selection(self):
        """Save current company selection to config."""
        selected = [company for company, var in self.company_vars.items() if var.get()]

        if not hasattr(self.config, 'company_filter'):
            from verificacion_correo.core.config import CompanyFilterConfig
            self.config.company_filter = CompanyFilterConfig()

        self.config.company_filter.enabled = self.company_filter_enabled.get()
        self.config.company_filter.companies = selected

        try:
            self.config.save()
        except Exception as e:
            logger.error(f"Error saving company filter config: {e}")

        self._update_company_status()

    def _update_company_status(self):
        """Update the company status label."""
        total = len(self.company_vars)
        selected = sum(1 for v in self.company_vars.values() if v.get())
        enabled = self.company_filter_enabled.get()

        if total == 0:
            self.company_status_label.config(text="No hay compañías cargadas", foreground='gray')
        elif not enabled:
            self.company_status_label.config(
                text=f"⏸️ Filtro deshabilitado — {total} compañías disponibles",
                foreground='gray'
            )
        else:
            self.company_status_label.config(
                text=f"✅ {total} compañías — {selected} seleccionadas para filtrar",
                foreground='#27ae60'
            )

    def _on_company_filter_toggle(self):
        """Handle company filter enable/disable toggle."""
        self._save_company_selection()

    def _get_selected_companies(self):
        """Get list of selected company names."""
        if not self.company_filter_enabled.get():
            return None
        selected = [company for company, var in self.company_vars.items() if var.get()]
        return selected if selected else None

    def _start_scraper(self):
        """Start the GAL scraper via OWA API."""
        if self.scraper_active:
            self._add_scraper_log("⚠️ El scraper ya está activo")
            return

        # Validate quantity
        if self.scraper_max_contacts.get() <= 0:
            messagebox.showwarning("Cantidad Inválida", "La cantidad debe ser mayor a 0")
            return

        session_file = self.config.get_session_file_path()
        if not Path(session_file).exists():
            messagebox.showwarning(
                "Sesión Requerida",
                "No hay sesión guardada.\nUse el botón 'Configurar Sesión' en la pestaña de Sesión."
            )
            return

        # Validate session health before starting
        self._add_scraper_log("🔍 Validando sesión antes de iniciar...")
        health = self.service.validate_session_api_quick()
        if not health.get('valid'):
            messagebox.showwarning(
                "Sesión Inválida",
                f"La sesión no es válida:\n{health.get('message', 'Error desconocido')}\n\n"
                "Por favor configure una nueva sesión antes de continuar."
            )
            return
        self._add_scraper_log(f"✅ Sesión válida (salud: {health.get('health', 'desconocido')})")

        # Confirm if resuming
        output_dir = Path(self.scraper_output_dir.get()) / "gal"
        excel_path = output_dir / "gal_directorio.xlsx"
        if excel_path.exists():
            resume = messagebox.askyesno(
                "Reanudar Extracción",
                f"Se encontró una extracción anterior:\n"
                f"{excel_path}\n\n"
                f"¿Desea continuar desde donde se quedó?\n\n"
                f"Seleccione 'No' para empezar desde cero (sobrescribirá)."
            )
            force_restart = not resume
        else:
            force_restart = True

        # Reset variables
        self.scraper_extracted_count.set(0)
        self._update_scraper_progress(0)
        self._update_scraper_total_label()

        # Update UI
        self.scraper_start_btn.config(state='disabled')
        self.scraper_stop_btn.config(state='normal')
        self.scraper_active = True

        max_contacts = self.scraper_max_contacts.get()
        self._add_scraper_log(f"🚀 Iniciando scraper (máx {max_contacts} contactos)...")
        self._update_scraper_status("🔄 Ejecutando...", "#3498db")

        # Get company filter
        company_filter = self._get_selected_companies()
        if company_filter:
            self._add_scraper_log(f"📋 Filtrando por {len(company_filter)} compañías")

        try:
            self.service.start_gal_scraping(
                excel_path=str(excel_path),
                max_contacts=max_contacts,
                force_restart=force_restart,
                company_filter=company_filter,
                address_list_id=self.selected_address_list_id.get(),
            )
        except Exception as e:
            self._add_scraper_log(f"❌ Error al iniciar: {e}")
            self.scraper_start_btn.config(state='normal')
            self.scraper_stop_btn.config(state='disabled')
            self.scraper_active = False

    def _stop_scraper(self):
        """Stop the scraper gracefully."""
        if not self.scraper_active:
            return

        if not messagebox.askyesno("Detener", "¿Estás seguro de que deseas detener la extracción?"):
            return

        self._add_scraper_log("⚠️ Deteniendo scraper...")
        self._update_scraper_status("⏸️ Deteniendo...", "#e67e22")
        self.service.stop_gal_scraping()
        self.scraper_stop_btn.config(state='disabled')

    def _start_enrichment(self):
        """Start enrichment of contacts from Excel."""
        excel_path = Path(self.scraper_output_dir.get()) / "gal" / "gal_directorio.xlsx"

        if not excel_path.exists():
            messagebox.showwarning("Error", "Primero ejecuta Extracción GAL")
            return

        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws2 = wb["Compañías"]
        companies_selected = False
        for row in range(2, ws2.max_row + 1):
            enrich = ws2.cell(row, 2).value
            if enrich and str(enrich).strip().upper() == 'X':
                companies_selected = True
                break

        if not companies_selected:
            messagebox.showinfo(
                "Sin selección",
                "Para completar información:\n\n"
                "1. Ve a la hoja 'Compañías' (segunda pestaña)\n"
                "2. Busca la empresa que quieres enriquecer\n"
                "3. Escribe 'X' en la columna 'Enrich'\n"
                "4. Guarda el archivo\n"
                "5. Vuelve a presionar 'Completar información'"
            )
            return

        self._add_scraper_log("🔄 Iniciando Completar información...")
        self._update_scraper_status("🔄 Completando...", "#3498db")

        try:
            self.service.start_enrichment(str(excel_path))
            self.scraper_start_btn.config(state='disabled')
            self.enrich_btn.config(state='disabled')
            self.scraper_active = True
        except Exception as e:
            self._add_scraper_log(f"❌ Error al iniciar enrichment: {e}")
            messagebox.showerror("Error", f"Error al iniciar enrichment: {e}")

    def _create_config_editor(self, parent):
        """Create configuration editor interface."""
        # Create notebook for different config sections
        config_notebook = ttk.Notebook(parent)
        config_notebook.pack(fill='both', expand=True)

        # Basic settings
        basic_frame = ttk.Frame(config_notebook)
        config_notebook.add(basic_frame, text="⚡ Básico")

        # URL
        ttk.Label(basic_frame, text="URL de OWA:").grid(row=0, column=0, sticky='w', padx=5, pady=5)
        self.owa_url_var = tk.StringVar(value=self.config.page_url)
        ttk.Entry(basic_frame, textvariable=self.owa_url_var, width=60).grid(row=0, column=1, padx=5, pady=5, sticky='ew')

        # Batch size
        ttk.Label(basic_frame, text="Tamaño de lote:").grid(row=1, column=0, sticky='w', padx=5, pady=5)
        self.batch_size_var = tk.IntVar(value=self.config.processing.batch_size)
        vcmd = (self.root.register(self._validate_numeric), '%P')
        batch_size_entry = ttk.Entry(basic_frame, textvariable=self.batch_size_var, width=10, validate='key', validatecommand=vcmd)
        batch_size_entry.grid(row=1, column=1, padx=5, pady=5, sticky='w')

        # Excel file
        ttk.Label(basic_frame, text="Archivo Excel:").grid(row=2, column=0, sticky='w', padx=5, pady=5)
        excel_frame = ttk.Frame(basic_frame)
        excel_frame.grid(row=2, column=1, padx=5, pady=5, sticky='ew')

        self.excel_file_var = tk.StringVar(value=self.config.get_excel_file_path())
        ttk.Entry(excel_frame, textvariable=self.excel_file_var, width=50).pack(side='left', padx=(0, 5))
        ttk.Button(excel_frame, text="Seleccionar", command=self._select_excel_config).pack(side='left')

        # Browser settings
        browser_frame = ttk.Frame(config_notebook)
        config_notebook.add(browser_frame, text="🌐 Navegador")

        # Headless mode
        self.headless_var = tk.BooleanVar(value=self.config.browser.headless)
        ttk.Checkbutton(browser_frame, text="Modo sin ventana (headless)", variable=self.headless_var).pack(anchor='w', padx=5, pady=5)

        # Discard draft option
        self.discard_draft_var = tk.BooleanVar(value=self.config.processing.discard_draft)
        ttk.Checkbutton(browser_frame, text="Descartar borrador automáticamente", variable=self.discard_draft_var).pack(anchor='w', padx=5, pady=5)

        # Session file
        ttk.Label(browser_frame, text="Archivo de sesión:").pack(anchor='w', padx=5, pady=(10, 0))
        session_frame = ttk.Frame(browser_frame)
        session_frame.pack(fill='x', padx=5, pady=5)

        self.session_file_var = tk.StringVar(value=self.config.get_session_file_path())
        ttk.Entry(session_frame, textvariable=self.session_file_var, width=50).pack(side='left', padx=(0, 5))
        ttk.Button(session_frame, text="Seleccionar", command=self._select_session_file).pack(side='left')

        # Default emails
        emails_frame = ttk.Frame(config_notebook)
        config_notebook.add(emails_frame, text="📧 Correos por Defecto")

        ttk.Label(emails_frame, text="Correos electrónicos de respaldo:").pack(anchor='w', padx=5, pady=5)

        # Text widget for emails
        self.emails_text = scrolledtext.ScrolledText(emails_frame, height=10, width=60)
        self.emails_text.pack(fill='both', expand=True, padx=5, pady=5)

        # Load current default emails
        default_emails_text = '\n'.join(self.config.default_emails)
        self.emails_text.insert('1.0', default_emails_text)

        # Configure grid weights
        basic_frame.columnconfigure(1, weight=1)

    def _select_excel_config(self):
        """Select Excel file for configuration."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo de Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_file_var.set(file_path)

    def _select_session_file(self):
        """Select session file for configuration."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo de sesión",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        if file_path:
            self.session_file_var.set(file_path)

    def _save_config(self):
        """Save configuration changes."""
        try:
            # Update configuration object
            self.config.page_url = self.owa_url_var.get()
            self.config.processing.batch_size = self.batch_size_var.get()
            self.config.processing.discard_draft = self.discard_draft_var.get()
            self.config.excel.default_file = self.excel_file_var.get()
            self.config.browser.headless = self.headless_var.get()
            self.config.browser.session_file = self.session_file_var.get()

            # Update default emails
            emails_text = self.emails_text.get('1.0', tk.END).strip()
            self.config.default_emails = [email.strip() for email in emails_text.split('\n') if email.strip()]

            # Save to file
            self.config.save()

            messagebox.showinfo("Configuración", "Configuración guardada exitosamente")
            self._refresh_excel_info()

        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar configuración: {e}")

    def _run_config_wizard(self):
        """Run configuration wizard for first-time setup."""
        wizard = ConfigWizard(self.root, self.config)
        if wizard.result:
            self._reload_config()

    def _create_status_bar(self, parent):
        """Create status bar."""
        status_frame = ttk.Frame(parent)
        status_frame.pack(fill='x', pady=(10, 0))

        # Status label
        self.status_label = ttk.Label(status_frame, text="Listo")
        self.status_label.pack(side='left')

        # Pending count label
        self.pending_label = ttk.Label(status_frame, text="", foreground='gray')
        self.pending_label.pack(side='left', padx=(20, 0))

        # Session status label
        self.session_indicator = ttk.Label(status_frame, text="", foreground='gray')
        self.session_indicator.pack(side='left', padx=(20, 0))

        # Session health indicator (API calls used)
        self.session_health_indicator = ttk.Label(status_frame, text="", foreground='gray')
        self.session_health_indicator.pack(side='left', padx=(20, 0))

        # Clock label
        self.clock_label = ttk.Label(status_frame)
        self.clock_label.pack(side='right')
        self._update_clock()

    def _setup_status_check(self):
        """Setup periodic status checks."""
        self._check_session_status()
        self._refresh_excel_info()
        self._check_progress()

    def _update_clock(self):
        """Update clock in status bar."""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.clock_label.config(text=now)
        self.root.after(1000, self._update_clock)

    def _check_progress(self):
        """Check for progress updates."""
        for item_type, data in self.service.check_queue():
            if item_type == 'complete':
                self._processing_complete(data)
            elif item_type == 'error':
                self._processing_error(data)
            elif item_type == 'api_complete':
                self._handle_api_complete(data)
            elif item_type == 'api_error':
                self._handle_api_error(data)
            elif item_type == 'gal_progress':
                self._update_gal_progress(data)
            elif item_type == 'gal_complete':
                self._handle_gal_complete(data)
            elif item_type == 'gal_error':
                self._handle_gal_error(data)
            elif item_type == 'company_scan_complete':
                self._handle_company_scan_complete(data)
            elif item_type == 'company_scan_error':
                self._handle_company_scan_error(data)
            elif item_type == 'address_list_scan_complete':
                self._handle_address_list_scan_complete(data)
            elif item_type == 'address_list_scan_error':
                self._handle_address_list_scan_error(data)
            elif item_type == 'session_health':
                self._update_session_health(data)
            elif item_type == 'progress':
                self._update_progress(data)
            elif item_type == 'enrich_progress':
                self._update_enrich_progress(data)
            elif item_type == 'enrich_complete':
                self._handle_enrich_complete(data)
            elif item_type == 'enrich_error':
                self._handle_enrich_error(data)

        if not self.service.is_processing:
            self.progress_bar.stop()
            if self.progress_bar['mode'] != 'determinate':
                self.progress_bar.config(mode='determinate')

        # Schedule next check
        self.root.after(100, self._check_progress)

    def _get_automation_engine_info(self) -> str:
        """Get information about the automation engine being used."""
        return "🤖 Motor: Playwright"

    def _select_excel_file(self):
        """Select Excel file dialog."""
        file_path = filedialog.askopenfilename(
            title="Seleccionar archivo de Excel",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_path_var.set(file_path)
            self._refresh_excel_info()

    def _refresh_excel_info(self):
        """Refresh Excel file information."""
        excel_path = self.excel_path_var.get()
        if not excel_path:
            return

        try:
            summary = self.service.get_excel_summary(excel_path)
            if 'error' in summary:
                self.summary_text.set(f"Error: {summary['error']}")
                self.start_btn.config(state='disabled')
            else:
                text = (f"Total de correos: {summary['total_emails']}\n"
                       f"Procesados: {summary['processed_count']}\n"
                       f"Pendientes: {summary['pending_count']}\n"
                       f"Lotes: {summary['batch_count']}")
                self.summary_text.set(text)

                # Enable/disable start button based on pending emails
                self.start_btn.config(
                    state='normal' if summary['pending_count'] > 0 else 'disabled'
                )

                # Update pending count in status bar
                self.pending_label.config(text=f"Pendientes: {summary['pending_count']}")

        except Exception as e:
            self.summary_text.set(f"Error: {e}")
            self.start_btn.config(state='disabled')

    def _start_processing(self):
        """Start email processing."""
        excel_path = self.excel_path_var.get()
        if not excel_path:
            messagebox.showwarning("Archivo Requerido", "Por favor seleccione un archivo de Excel")
            return

        # Validate session first
        session_status = self.service.validate_session()
        if not session_status.get('is_valid'):
            if not messagebox.askyesno(
                "Sesión Inválida",
                "La sesión del navegador no es válida o ha expirado.\n¿Desea configurar una nueva sesión?"
            ):
                return
            self._setup_session()
            return

        # Confirm processing
        if not messagebox.askyesno(
            "Confirmar Procesamiento",
            "¿Desea iniciar el procesamiento de correos pendientes?"
        ):
            return

        try:
            self.service.start_processing(excel_path)
            self.start_btn.config(state='disabled')
            self.stop_btn.config(state='normal')
            self.progress_text.set("Iniciando procesamiento...")
            self._add_log("🚀 Iniciando procesamiento de correos")
            self.status_label.config(text="Procesando...")
            self.is_processing = True

        except Exception as e:
            messagebox.showerror("Error", f"Error al iniciar procesamiento: {e}")

    def _stop_processing(self):
        """Stop email processing."""
        if not messagebox.askyesno("Detener", "¿Estás seguro de que deseas detener el procesamiento?"):
            return
        try:
            self.service.stop_processing()
            self.start_btn.config(state='normal')
            self.stop_btn.config(state='disabled')
            self.progress_text.set("Procesamiento detenido")
            self._add_log("⏹ Procesamiento detenido por el usuario")
            self.status_label.config(text="Detenido")

        except Exception as e:
            messagebox.showerror("Error", f"Error al detener procesamiento: {e}")

    def _processing_complete(self, stats):
        """Handle processing completion."""
        self.root.after(0, self._handle_processing_complete, stats)

    def _handle_processing_complete(self, stats):
        """Handle processing completion in GUI thread."""
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.progress_text.set("Procesamiento completado")
        self.status_label.config(text="Completado")
        self.is_processing = False

        # Show results
        total = stats.total_emails
        success = stats.successful
        not_found = stats.not_found
        errors = stats.errors

        message = f"""Procesamiento completado:

📧 Total procesados: {total}
✅ Exitosos: {success} ({success/total*100:.1f}%)
❌ No encontrados: {not_found} ({not_found/total*100:.1f}%)
⚠️ Errores: {errors} ({errors/total*100:.1f}%)

Duración: {stats.duration_seconds:.1f} segundos
Resultados guardados en: {self.excel_path_var.get()}"""

        messagebox.showinfo("Procesamiento Completado", message)
        self._add_log(f"✅ Procesamiento completado: {success} exitosos, {not_found} no encontrados, {errors} errores")
        self._refresh_excel_info()
        self._refresh_results_tree()
        self._save_run_history("playwright", {
            "total": total, "success": success, "not_found": not_found,
            "errors": errors, "duration": stats.duration_seconds
        })

    def _processing_error(self, error_msg):
        """Handle processing error."""
        self.root.after(0, self._handle_processing_error, error_msg)

    def _handle_processing_error(self, error_msg):
        """Handle processing error in GUI thread."""
        self.start_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.progress_text.set("Error en procesamiento")
        self.status_label.config(text="Error")
        self.is_processing = False

        messagebox.showerror("Error de Procesamiento", f"Ocurrió un error:\n{error_msg}")
        self._add_log(f"❌ Error de procesamiento: {error_msg}")

    def _start_api_search(self):
        """Start API-based contact search."""
        excel_path = self.excel_path_var.get()
        if not excel_path:
            messagebox.showwarning("Archivo Requerido", "Por favor seleccione un archivo de Excel")
            return

        session_file = self.config.get_session_file_path()
        if not Path(session_file).exists():
            if not messagebox.askyesno(
                "Sesión Requerida",
                "El archivo de sesión no existe.\n¿Desea configurar una nueva sesión primero?"
            ):
                return
            self._setup_session()
            return

        # Validate session health before starting
        self._add_log("🔍 Validando sesión antes de iniciar...")
        health = self.service.validate_session_api_quick()
        if not health.get('valid'):
            messagebox.showwarning(
                "Sesión Inválida",
                f"La sesión no es válida:\n{health.get('message', 'Error desconocido')}\n\n"
                "Por favor configure una nueva sesión antes de continuar."
            )
            return

        # Confirm
        if not messagebox.askyesno(
            "Confirmar Búsqueda por API",
            "¿Desea buscar contactos en el directorio de OWA vía API?\n\n"
            "Esto consultará directamente el servicio FindPeople de Exchange\n"
            "usando las cookies de la sesión guardada."
        ):
            return

        try:
            self.service.start_api_processing(excel_path)
            self.start_btn.config(state='disabled')
            self.api_btn.config(state='disabled')
            self.stop_btn.config(state='normal')
            self.progress_text.set("Buscando contactos por API...")
            self._add_log("🔍 Iniciando búsqueda de contactos vía API de OWA")
            self.status_label.config(text="Buscando...")
            self.is_processing = True

        except Exception as e:
            messagebox.showerror("Error", f"Error al iniciar búsqueda API: {e}")

    def _handle_api_complete(self, result):
        """Handle API search completion in GUI thread."""
        self.start_btn.config(state='normal')
        self.api_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.progress_var.set(0)
        self.progress_bar['value'] = 0
        self.is_processing = False

        total = result.get("total", 0)
        success = result.get("success", 0)
        not_found = result.get("not_found", 0)
        errors = result.get("errors", 0)
        duration = result.get("duration", 0)
        session_expired = result.get("expired", False)
        found_pct = (success / total * 100) if total > 0 else 0

        if session_expired:
            remaining = result.get("remaining", 0)
            self.progress_text.set(f"⚠️ Sesión expirada tras {total} consultas")
            self.status_label.config(text="Sesión expirada")
            message = f"""Búsqueda por API DETENIDA - Sesión expirada

⚠️ La sesión de OWA ha expirado después de {total} consultas.
   Los emails restantes quedarán pendientes para el próximo ciclo.

📧 Procesados: {total}
✅ Encontrados: {success} ({found_pct:.1f}%)
❌ No encontrados: {not_found}
⚠️ Errores (sesión): {errors}

            📌 Para continuar:
                1. Cierra sesión y vuelve a iniciar en OWA
                2. Usa 'Configurar Sesión' en la pestaña de Sesión
                3. Vuelve a ejecutar la búsqueda

Duración: {duration:.1f} segundos"""

            messagebox.showwarning("Sesión Expirada", message)
            self._add_log(f"⚠️ API search: sesión expirada tras {total} consultas ({success} encontrados, {not_found} no encontrados)")
            self.notebook.select(2)
        else:
            self.progress_text.set("Búsqueda API completada")
            self.status_label.config(text="Completado")
            message = f"""Búsqueda por API completada:

📧 Total procesados: {total}
✅ Encontrados: {success} ({found_pct:.1f}%)
❌ No encontrados: {not_found}
⚠️ Errores: {errors}

Duración: {duration:.1f} segundos
Resultados guardados en: {self.excel_path_var.get()}"""

            messagebox.showinfo("Búsqueda por API Completada", message)
            self._add_log(f"✅ API search: {success} encontrados, {not_found} no encontrados, {errors} errores")

        self._refresh_excel_info()
        self._refresh_results_tree()
        self._save_run_history("api", {
            "total": total, "success": success, "not_found": not_found,
            "errors": errors, "duration": duration
        })

    def _handle_api_error(self, error_msg):
        """Handle API search error in GUI thread."""
        self.start_btn.config(state='normal')
        self.api_btn.config(state='normal')
        self.stop_btn.config(state='disabled')
        self.progress_var.set(0)
        self.progress_bar['value'] = 0
        self.is_processing = False
        self.progress_text.set("Error en búsqueda API")
        self.status_label.config(text="Error")

        messagebox.showerror("Error de Búsqueda API", f"Ocurrió un error:\n{error_msg}")
        self._add_log(f"❌ Error en búsqueda API: {error_msg}")

    def _update_gal_progress(self, data):
        """Update GAL scraper progress from background thread."""
        count = data.get("count", 0)
        total = data.get("total", 0) or 1
        self.scraper_extracted_count.set(count)
        self._update_scraper_progress(count)
        if total > 0:
            pct = (count / total) * 100
            self.scraper_progress_bar["value"] = pct

    def _handle_gal_complete(self, result):
        """Handle GAL scraper completion."""
        self.scraper_start_btn.config(state="normal")
        self.scraper_stop_btn.config(state="disabled")
        self.scraper_active = False

        total = result.get("total", 0)
        total_scanned = result.get("total_scanned", 0)
        duration = result.get("duration", 0)
        expired = result.get("expired", False)
        stopped = result.get("stopped", False)
        files = result.get("files", {})
        filtered_companies = result.get("filtered_companies")

        self.scraper_extracted_count.set(total)
        self._update_scraper_progress(total)

        if expired:
            self._update_scraper_status("⚠️ Sesión expirada", "#e67e22")
            self._add_scraper_log(f"⚠️ Sesión expirada tras {total} contactos")
            messagebox.showwarning(
                "Sesión Expirada",
                f"La sesión expiró después de {total} contactos.\n\n"
                f"El progreso se ha guardado. Para continuar:\n"
                f"1. Vuelva a iniciar sesión en OWA\n"
                f"2. Use 'Configurar Sesión' en la pestaña de Sesión\n"
                f"3. Inicie la extracción nuevamente (se reanudará)\n\n"
                f"📁 {files.get('json', '')}"
            )
            self.notebook.select(2)
        elif stopped:
            self._update_scraper_status("⏹️ Detenido", "#e67e22")
            self._add_scraper_log(f"⏹️ Scraper detenido ({total} contactos)")
            msg = f"Scraper detenido con {total} contactos guardados."
            if filtered_companies:
                msg += f"\n\nFiltrado por: {', '.join(filtered_companies)}"
            msg += f"\n\nResultados guardados en:\n  {files.get('json', '')}\n  {files.get('csv', '')}"
            messagebox.showinfo("Extracción Detenida", msg)
        else:
            self._update_scraper_status("✅ Completado", "#27ae60")
            self._add_scraper_log(f"✅ Extracción completada: {total} contactos en {duration:.1f}s")
            msg = f"Directorio extraído: {total} contactos"
            if filtered_companies:
                msg += f"\nFiltrado por: {', '.join(filtered_companies)}"
            msg += f"\n\nEscaneados: {total_scanned} | Guardados: {total}"
            msg += f"\nDuración: {duration:.1f}s"
            msg += f"\n\nResultados guardados en:\n  {files.get('json', '')}\n  {files.get('csv', '')}"
            messagebox.showinfo("Extracción Completada", msg)

        self._save_run_history("gal", {
            "total": total, "duration": duration, "expired": expired,
            "stopped": stopped, "files": files
        })

    def _handle_gal_error(self, error_msg):
        """Handle GAL scraper error."""
        self.scraper_start_btn.config(state="normal")
        self.scraper_stop_btn.config(state="disabled")
        self.scraper_active = False
        self._update_scraper_status("❌ Error", "#e74c3c")
        self._add_scraper_log(f"❌ Error: {error_msg}")
        messagebox.showerror("Error de Extracción", f"Ocurrió un error:\n{error_msg}")

    def _update_enrich_progress(self, data):
        """Update enrichment progress."""
        count = data.get('count', 0)
        companies = data.get('companies', 0)
        self._update_scraper_status(f"🔄 Enriqueciendo: {count} contactos", "#3498db")

    def _handle_enrich_complete(self, data):
        """Handle enrichment completion."""
        self.scraper_start_btn.config(state="normal")
        self.enrich_btn.config(state="normal")
        self.scraper_stop_btn.config(state="disabled")
        self.scraper_active = False

        error = data.get('error')
        if error:
            self._update_scraper_status("⚠️ Sin compañías", "#e67e22")
            self._add_scraper_log(f"⚠️ {error}")
            messagebox.showwarning("Enrichment", error)
            return

        contacts = data.get('contacts_enriched', 0)
        companies = data.get('companies_done', 0)

        self._update_scraper_status("✅ Completado", "#27ae60")
        self._add_scraper_log(f"✅ Completado: {contacts} contactos de {companies} compañías")
        messagebox.showinfo(
            "Completado",
            f"Se completaron {contacts} contactos de {companies} compañías.\n\n"
            f"Revisa el Excel para ver los datos."
        )

    def _handle_enrich_error(self, error_msg):
        """Handle enrichment error."""
        self.scraper_start_btn.config(state="normal")
        self.enrich_btn.config(state="normal")
        self.scraper_stop_btn.config(state="disabled")
        self.scraper_active = False
        self._update_scraper_status("❌ Error enrichment", "#e74c3c")
        self._add_scraper_log(f"❌ Error enrichment: {error_msg}")
        messagebox.showerror("Error de Enrichment", f"Ocurrió un error:\n{error_msg}")

    MAX_LOG_MESSAGES = 1000

    def _add_log(self, message):
        """Add message to log."""
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_entry = f"[{timestamp}] {message}\n"
        self.log_messages.append(log_entry)
        if len(self.log_messages) > self.MAX_LOG_MESSAGES:
            self.log_messages.pop(0)
            self.log_text.delete('1.0', '2.0')

        self.log_text.insert(tk.END, log_entry)
        self.log_text.see(tk.END)

    def _clear_log(self):
        """Clear log messages."""
        self.log_text.delete('1.0', tk.END)
        self.log_messages.clear()

    def _save_log(self):
        """Save log messages to file."""
        file_path = filedialog.asksaveasfilename(
            title="Guardar Log",
            defaultextension=".log",
            filetypes=[("Log files", "*.log"), ("Text files", "*.txt"), ("All files", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.writelines(self.log_messages)
                messagebox.showinfo("Log Guardado", f"Log guardado en:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Error", f"Error al guardar log: {e}")

    def _check_session_status(self):
        """Check and display session status with real API validation."""
        try:
            session_file = Path(self.config.get_session_file_path())
            session_type = "Playwright (Chromium)"

            status_text = f"Tipo: {session_type}\n"
            status_text += f"Archivo: {session_file}\n"
            status_text += f"Existe: {'Sí' if session_file.exists() else 'No'}\n"

            is_valid = False

            if session_file.exists():
                try:
                    with open(session_file, 'r') as f:
                        session_data = json.load(f)
                    cookies_count = len(session_data.get('cookies', []))

                    # Validate session via API
                    api_result = self.service.validate_session_api_quick()
                    is_valid = api_result.get('valid', False)
                    health = api_result.get('health', 'unknown')
                    message = api_result.get('message', '')

                    if is_valid:
                        status_text += f"✅ Sesión válida (salud: {health})\n"
                        status_text += f"Cookies: {cookies_count}\n"
                        if message:
                            status_text += f"Info: {message}"
                    else:
                        status_text += f"❌ Sesión inválida\n"
                        status_text += f"Cookies: {cookies_count}\n"
                        if message:
                            status_text += f"Razón: {message}"
                        else:
                            status_text += f"\n⚠️ La sesión ha expirado o es inválida\n"
                            status_text += f"Haz clic en 'Iniciar Sesión' para autenticarte"

                except (json.JSONDecodeError, KeyError) as e:
                    status_text += f"❌ Error al leer sesión: {e}"
            else:
                status_text += f"❌ No existe sesión\n"
                status_text += f"\n⚠️ Sesión no encontrada\n"
                status_text += f"Haz clic en 'Iniciar Sesión' para crear una"

            self.session_status_text.set(status_text)

            # Update session indicator in status bar
            if session_file.exists() and is_valid:
                self.session_indicator.config(text="🟢 Sesión OK", foreground='#27ae60')
            else:
                self.session_indicator.config(text="🔴 Sin sesión", foreground='#e74c3c')

            # Update action button based on session validity
            if is_valid:
                self.session_action_btn.config(text="🔄 Verificar Sesión")
                ToolTip(self.session_action_btn, "Verifica que la sesión sigue siendo válida")
            else:
                self.session_action_btn.config(text="🔐 Iniciar Sesión")
                ToolTip(self.session_action_btn, "Abre navegador para iniciar sesión en OWA")

            # Update session info text
            self.session_info_text.config(state='normal')
            self.session_info_text.delete('1.0', tk.END)
            self.session_info_text.insert('1.0', f"Información de la sesión:\n\n{status_text}")
            self.session_info_text.config(state='disabled')

        except Exception as e:
            self.session_status_text.set(f"Error: {e}")

    def _update_session_health(self, health_info):
        """Update session health indicator in status bar."""
        calls_used = health_info.get('calls_used', 0)
        estimated_limit = health_info.get('estimated_limit', 40)
        health = health_info.get('health', 'unknown')
        message = health_info.get('message', '')

        # Determine color and icon based on health
        if health == 'expired':
            color = '#e74c3c'
            icon = '🔴'
            text = f"{icon} Sesión EXPIRADA"
        elif health == 'danger':
            color = '#e74c3c'
            icon = '🔴'
            text = f"{icon} Sesión inválida"
        elif health == 'warning':
            color = '#e67e22'
            icon = '🟡'
            text = f"{icon} API: {calls_used}/{estimated_limit}"
        elif health == 'ok':
            # Show warning color when approaching limit
            if calls_used > 0 and calls_used >= estimated_limit * 0.7:
                color = '#e67e22'
                icon = '🟡'
            else:
                color = '#27ae60'
                icon = '🟢'
            text = f"{icon} API: {calls_used}/{estimated_limit}"
        else:
            color = 'gray'
            icon = '⚪'
            text = f"{icon} API: ?/?"

        self.session_health_indicator.config(text=text, foreground=color)

    def _handle_session_action(self):
        """Handle the dynamic session action button click."""
        session_file = Path(self.config.get_session_file_path())
        if session_file.exists():
            # Check if session is valid
            api_result = self.service.validate_session_api_quick()
            if api_result.get('valid'):
                # Session is valid - just verify and show status
                self._check_session_status()
            else:
                # Session exists but is invalid - need to setup new session
                self._setup_session()
        else:
            # No session exists - setup new session
            self._setup_session()

    def _setup_session(self):
        """Set up browser session."""
        browser_type = "Chromium con Playwright"
        session_file = "state.json"

        if not messagebox.askyesno(
            "Configurar Sesión",
            f"Se abrirá {browser_type} para que inicie sesión manualmente.\n"
            f"La sesión se guardará en: {session_file}\n\n"
            "Después de iniciar sesión, vuelve a esta ventana.\n\n"
            "¿Desea continuar?"
        ):
            return

        try:
            # Use Playwright session setup
            self._setup_playwright_session()

        except Exception as e:
            messagebox.showerror("Error", f"Error al iniciar configuración de sesión: {e}")

    def _setup_playwright_session(self):
        """Set up Playwright (Chromium) session."""
        self._session_confirm_closing = False
        
        def setup_in_background():
            try:
                success = self.service.setup_session()
                
                # Schedule GUI cleanup on main thread (never access GUI directly)
                self.root.after(0, lambda: self._cleanup_session_after_setup(success))
                
            except Exception as e:
                self.root.after(0, lambda: self._cleanup_session_after_setup(False, str(e)))

        setup_thread = threading.Thread(target=setup_in_background, daemon=True)
        setup_thread.start()

        # Show confirmation window with "Sesión Lista" button
        self._show_session_confirm_dialog()

    def _show_session_confirm_dialog(self):
        """Show a dialog with 'Sesión Lista' button for manual confirmation."""
        import tkinter as tk
        from tkinter import ttk
        
        # Create a new top-level window
        confirm_win = tk.Toplevel(self.root)
        confirm_win.title("Configurando Sesión...")
        confirm_win.geometry("350x200")
        confirm_win.resizable(False, False)
        confirm_win.transient(self.root)
        confirm_win.grab_set()
        
        # Store reference for cleanup (before any scheduling)
        self._session_confirm_window = confirm_win
        self._session_confirm_closing = False
        
        # Center the window using after to avoid update_idletasks issues on macOS
        def _position_window():
            try:
                if confirm_win.winfo_exists():
                    sw = confirm_win.winfo_screenwidth()
                    sh = confirm_win.winfo_screenheight()
                    x = (sw // 2) - (350 // 2)
                    y = (sh // 2) - (200 // 2)
                    confirm_win.geometry(f"350x200+{x}+{y}")
            except tk.TclError:
                pass
        self.root.after(50, _position_window)
        
        # Instructions
        instructions = ttk.Label(
            confirm_win,
            text="1. Inicia sesión en el navegador\n"
                 "2. Espera a que cargue tu bandeja\n"
                 "3. Haz clic en 'Sesión Lista'",
            justify='left',
            font=('TkDefaultFont', 10)
        )
        instructions.pack(padx=20, pady=(20, 10), anchor='w')
        
        # Status label
        self._confirm_status = ttk.Label(
            confirm_win,
            text="⏳ Esperando...",
            foreground='gray'
        )
        self._confirm_status.pack(pady=5)
        
        # Confirm button
        confirm_btn = ttk.Button(
            confirm_win,
            text="✅ Sesión Lista",
            command=self._on_session_confirmed,
            style='Accent.TButton'
        )
        confirm_btn.pack(pady=10, ipadx=20, ipady=5)
        
        # Make window closeable
        confirm_win.protocol("WM_DELETE_WINDOW", lambda: self._on_session_confirm_closed())
    
    def _on_session_confirmed(self):
        """Handle 'Sesión Lista' button click."""
        if self._session_confirm_closing:
            return
        try:
            self.service.confirm_session_ready()
            if self._confirm_status and self._confirm_status.winfo_exists():
                self._confirm_status.config(text="✅ Sesión confirmada, guardando...", foreground='green')
            # Close the confirm window after a short delay
            self.root.after(1500, lambda: self._close_session_confirm())
        except Exception as e:
            logger.error(f"Error confirming session: {e}")
    
    def _on_session_confirm_closed(self):
        """Handle confirm window close button."""
        self._close_session_confirm()
    
    def _close_session_confirm(self):
        """Close the session confirmation window safely."""
        if self._session_confirm_closing:
            return
        self._session_confirm_closing = True
        if self._session_confirm_window:
            try:
                if self._session_confirm_window.winfo_exists():
                    self._session_confirm_window.grab_release()
            except tk.TclError:
                pass
            try:
                if self._session_confirm_window.winfo_exists():
                    self._session_confirm_window.destroy()
            except tk.TclError:
                pass
            self._session_confirm_window = None
        self._confirm_status = None
    
    def _cleanup_session_after_setup(self, success: bool, error_msg: str = ""):
        """Clean up after session setup completes (always runs on main thread)."""
        self._close_session_confirm()
        
        if success:
            messagebox.showinfo(
                "Éxito",
                "Sesión configurada correctamente.\n\n"
                "La sesión ha sido guardada en state.json"
            )
        elif error_msg:
            messagebox.showerror(
                "Error",
                f"Error durante la configuración de sesión: {error_msg}"
            )
        else:
            messagebox.showwarning(
                "Sesión No Detectada",
                "No se pudo detectar la sesión automáticamente.\n\n"
                "Si ya iniciaste sesión, vuelve a intentar.\n"
                "Si no pudiste acceder, verifica tus credenciales."
            )
        
        self._check_session_status()

    def _delete_session(self):
        """Delete browser session."""
        if not messagebox.askyesno(
            "Eliminar Sesión",
            "¿Está seguro de que desea eliminar la sesión guardada?\n"
            "Necesitará configurar una nueva sesión para continuar."
        ):
            return

        try:
            session_file = Path(self.config.get_session_file_path())
            if session_file.exists():
                session_file.unlink()
                self._add_log("🗑️ Sesión eliminada correctamente")
                messagebox.showinfo(
                    "Sesión Eliminada",
                    f"La sesión ha sido eliminada:\n{session_file}"
                )
            else:
                messagebox.showinfo(
                    "Sin Sesión",
                    "No hay archivo de sesión para eliminar."
                )
            self._check_session_status()

        except Exception as e:
            messagebox.showerror("Error", f"Error al eliminar sesión: {e}")

    def _open_excel_file(self):
        """Open Excel file with system default application."""
        excel_path = self.excel_path_var.get()
        if not excel_path or not Path(excel_path).exists():
            messagebox.showwarning("Archivo no encontrado", "El archivo de Excel no existe")
            return

        if not open_file(excel_path):
            messagebox.showerror("Error", "No se pudo abrir el archivo Excel")

    def _open_data_folder(self):
        """Open data folder in system file explorer."""
        data_path = Path(self.config.get_excel_file_path()).parent
        if not open_folder(data_path):
            messagebox.showerror("Error", "No se pudo abrir la carpeta de datos")

    def _refresh_results_tree(self):
        """Refresh results treeview from Excel file."""
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        try:
            excel_path = self.excel_path_var.get()
            if not excel_path or not Path(excel_path).exists():
                return
            from verificacion_correo.core.excel import ProcessingStatus
            reader = ExcelReader(excel_path)
            records = reader.read_all_records()
            for rec in records[-50:]:
                status_char = '✅' if rec.status == ProcessingStatus.SUCCESS else ('❌' if rec.status == ProcessingStatus.ERROR else '⬜')
                status_label = rec.status.value if hasattr(rec.status, 'value') else str(rec.status)
                nombre = (rec.data or {}).get('name', '')
                email_personal = (rec.data or {}).get('email', '')
                telefono = (rec.data or {}).get('phone', '')
                self.results_tree.insert('', 'end', values=(
                    rec.email, f"{status_char} {status_label}",
                    nombre[:40], email_personal[:40], telefono[:30]
                ))
        except Exception as e:
            self._add_log(f"⚠️ No se pudieron cargar resultados: {e}")

    def _reload_config(self):
        """Reload configuration."""
        try:
            from ..core.config import reload_config
            reload_config()
            self.config = Config()
            messagebox.showinfo("Configuración", "Configuración recargada exitosamente")
            self._refresh_excel_info()
            self._check_session_status()
        except Exception as e:
            messagebox.showerror("Error", f"Error al recargar configuración: {e}")

    def _setup_keyboard_shortcuts(self):
        """Setup keyboard shortcuts."""
        self.root.bind('<Control-Return>', lambda e: self._start_processing() if not self.is_processing else None)
        self.root.bind('<Escape>', lambda e: self._stop_processing() if self.is_processing else None)
        self.root.bind('<Control-s>', lambda e: self._save_config())
        self.root.bind('<Control-Shift-Return>', lambda e: self._start_scraper())

    def _setup_safe_close(self):
        """Setup safe window close handler to prevent Tkinter crashes on macOS."""
        self.root.protocol("WM_DELETE_WINDOW", self._safe_close)

    def _safe_close(self):
        """Safely close the application, cleaning up all windows."""
        # Close session confirm window first if it exists
        self._close_session_confirm()
        # Then quit and destroy main window
        try:
            self.root.quit()
        except tk.TclError:
            pass
        try:
            self.root.destroy()
        except tk.TclError:
            pass

    def _validate_numeric(self, value_if_allowed):
        """Validate numeric input for Entry/Spinbox."""
        if value_if_allowed == "":
            return True
        try:
            int(value_if_allowed)
            return True
        except ValueError:
            return False

    def _show_about(self):
        """Show About dialog."""
        from verificacion_correo import __version__
        messagebox.showinfo(
            "Acerca de Verificación de Correos",
            "Verificación de Correos OWA\n"
            "Versión: {}\n\n"
            "Herramienta de automatización para extracción de contactos desde OWA Madrid.\n\n"
            "Atajos:\n"
            "  Ctrl+Enter        - Iniciar procesamiento\n"
            "  Escape            - Detener procesamiento\n"
            "  Ctrl+S            - Guardar configuración\n"
            "  Ctrl+Shift+Enter  - Iniciar extracción GAL".format(__version__)
        )

    def _open_scraper_output(self):
        """Open scraper output directory in file explorer."""
        output_dir = Path(self.config.get_excel_file_path()).parent / "gal"
        if output_dir.exists():
            if not open_folder(output_dir):
                messagebox.showerror("Error", "No se pudo abrir la carpeta de resultados")
        else:
            messagebox.showinfo("Sin resultados", "No hay resultados de extracción aún. Ejecuta el scraper primero.")

    def _on_tab_changed(self, event=None):
        """Handle notebook tab change."""
        tab_id = self.notebook.select()
        if tab_id:
            tab_index = self.notebook.index(tab_id)
            tab_names = ["Procesamiento", "Sesión", "Configuración", "Scraper"]
            if tab_index < len(tab_names):
                self._update_status("Pestaña: {}".format(tab_names[tab_index]))

    def _update_status(self, message):
        """Update status bar message."""
        self.status_label.config(text=message)

    def _save_run_history(self, run_type, stats):
        """Save run history to JSON file."""
        history_file = Path(self.config.get_excel_file_path()).parent / "run_history.json"
        history = []
        if history_file.exists():
            try:
                with open(history_file, encoding='utf-8') as f:
                    history = json.load(f)
            except (json.JSONDecodeError, IOError):
                pass
        history.append({
            "type": run_type,
            "timestamp": datetime.now().isoformat(),
            "stats": stats
        })
        history = history[-20:]
        try:
            with open(history_file, 'w', encoding='utf-8') as f:
                json.dump(history, f, indent=2, default=str)
        except IOError:
            pass

    def _update_progress(self, progress_data):
        """Update progress bar with current/total info."""
        current = progress_data.get('current', 0)
        total = progress_data.get('total', 0) or 1
        pct = (current / total) * 100
        self.progress_var.set(pct)
        self.progress_bar.config(mode='determinate')
        self.progress_bar['value'] = pct
        self.progress_text.set(f"Procesando: {current}/{total} ({pct:.0f}%)")


def main():
    """Main entry point for GUI."""
    root = tk.Tk()

    # Set up modern styling if available
    try:
        style = ttk.Style()
        if 'clam' in style.theme_names():
            style.theme_use('clam')
        style.configure('Accent.TButton', font=('TkDefaultFont', 10, 'bold'))
        style.configure('Treeview', rowheight=28)
    except tk.TclError:
        pass

    app = VerificacionCorreosGUI(root)
    root.mainloop()


if __name__ == "__main__":
    main()
