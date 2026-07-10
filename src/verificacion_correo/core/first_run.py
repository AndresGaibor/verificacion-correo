"""
First run setup for verificacion-correo.

This module handles the initial setup when the application runs for the first time,
including creating default configuration, detecting missing resources, and
guiding users through initial configuration.
"""

import os
import sys
import shutil
import subprocess
from pathlib import Path
from typing import Optional, Dict, Any
import logging

from verificacion_correo.core.config import Config


logger = logging.getLogger(__name__)


class FirstRunManager:
    """Manages first-time setup and initialization."""

    def __init__(self):
        """Initialize first run manager."""
        self.config = None

    def _get_marker_path(self) -> Path:
        """Get path for first run marker file."""
        if self.config:
            data_dir = Path(self.config.get_excel_file_path()).parent
            return data_dir / ".first_run_completed"
        return Path(".first_run_completed")

    def is_first_run(self) -> bool:
        """Check if this is the first run."""
        # Check for essential configuration files
        config_paths = [
            Path("config.yaml"),
            Path("config/default.yaml")
        ]

        for config_path in config_paths:
            if config_path.exists():
                return False

        # Check first run marker
        marker = self._get_marker_path()
        if marker.exists():
            return False

        return True

    def run_first_time_setup(self, config_path: Optional[str] = None) -> Config:
        """
        Run first-time setup process.

        Args:
            config_path: Optional path to configuration file

        Returns:
            Initialized Config object
        """
        print("🚀 Iniciando configuración inicial de Verificación de Correos OWA")
        print("=" * 60)

        try:
            # Step 1: Create configuration
            print("📋 Paso 1: Creando configuración...")
            self.config = Config(config_path)
            print(f"✅ Configuración creada: {self.config._config_path}")

            # Step 2: Ensure data directory exists
            print("📁 Paso 2: Verificando directorios...")
            self._ensure_directories()
            print("✅ Directorios verificados")

            # Step 3: Check/create Excel file
            print("📊 Paso 3: Verificando archivo de Excel...")
            excel_status = self._ensure_excel_file()
            print(f"✅ Archivo Excel: {excel_status}")

            # Step 4: Check and install Playwright browsers
            print("🌐 Paso 4: Verificando navegadores de Playwright...")
            browser_status = self._check_and_install_playwright_browsers()
            print(f"📌 Navegadores: {browser_status}")

            # Step 5: Check session status
            print("🔐 Paso 5: Verificando sesión del navegador...")
            session_status = self._check_session_status()
            print(f"📌 Sesión: {session_status}")

            # Step 6: Create first run marker
            print("🎯 Paso 6: Finalizando configuración...")
            self._create_first_run_marker()
            print("✅ Configuración inicial completada")

            print("\n" + "=" * 60)
            print("🎉 ¡Listo para usar! La aplicación está configurada.")
            print("\nPróximos pasos:")
            print("1. Configura la sesión del navegador usando la GUI o el comando CLI")
            print("2. Abre el archivo Excel y añade los correos que quieres procesar")
            print("3. Inicia el procesamiento desde la interfaz gráfica")

            return self.config

        except Exception as e:
            logger.error(f"Error en configuración inicial: {e}")
            print(f"\n❌ Error en configuración inicial: {e}")
            print("Por favor, revisa los archivos y vuelve a intentar.")
            raise

    def _ensure_directories(self):
        """Ensure necessary directories exist."""
        directories = [
            Path("data"),
            Path("logs"),
            Path("sessions"),
            Path("exports")
        ]

        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
            logger.debug(f"Directory ensured: {directory}")

    def _ensure_excel_file(self) -> str:
        """Ensure Excel file exists with proper structure."""
        excel_path = Path(self.config.get_excel_file_path())

        if excel_path.exists():
            return f"Existente en {excel_path}"
        else:
            # Try to copy from template
            template_paths = [
                Path("data/correos_template.xlsx"),
                Path("data/correos.xlsx"),
            ]

            for template_path in template_paths:
                if template_path.exists():
                    shutil.copy(template_path, excel_path)
                    return f"Creado desde plantilla {template_path}"

            # Create basic Excel file
            return self._create_basic_excel_file(excel_path)

    def _create_basic_excel_file(self, excel_path: Path) -> str:
        """Create a basic Excel file with sample data."""
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active
            ws.title = "Contactos"

            # Headers
            headers = [
                "Correo", "Estado", "Nombre", "Email Personal", "Teléfono",
                "Dirección", "Departamento", "Compañía", "Oficina", "SIP",
                "Fecha Procesamiento", "Observaciones"
            ]

            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Sample data
            sample_emails = [
                "ASP164@MADRID.ORG",
                "AGM564@MADRID.ORG",
                "USR789@MADRID.ORG"
            ]

            for row_idx, email in enumerate(sample_emails, start=2):
                ws.cell(row=row_idx, column=1, value=email)  # Email
                ws.cell(row=row_idx, column=2, value="")     # Status (empty = pending)

            # Adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20

            wb.save(excel_path)
            return f"Creado con datos de ejemplo en {excel_path}"

        except ImportError:
            return "No se pudo crear (openpyxl no disponible)"
        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            return f"Error al crear: {e}"

    def _check_session_status(self) -> str:
        """Check browser session status."""
        session_file = Path(self.config.get_session_file_path())

        if session_file.exists():
            size = session_file.stat().st_size
            if size > 100:  # Arbitrary threshold for "reasonable" session file
                return f"Existente ({size} bytes) - probablemente válida"
            else:
                return f"Existente pero pequeño ({size} bytes) - puede ser inválida"
        else:
            return f"No existe (esperado en {session_file})"

    def _check_and_install_playwright_browsers(self) -> str:
        """
        Check if Playwright browsers are installed and install them if needed.

        Returns:
            Status message about browser installation
        """
        try:
            # Try to check if Chromium is available
            from playwright.sync_api import sync_playwright

            try:
                with sync_playwright() as p:
                    # Try to get browser executable path
                    # This will fail if browsers are not installed
                    browser_path = p.chromium.executable_path
                    if browser_path and Path(browser_path).exists():
                        return f"Navegadores ya instalados en {browser_path}"
            except Exception:
                # Browsers not installed, need to install
                pass

            print("📥 Descargando navegadores de Playwright...")
            print("   Esto puede tomar varios minutos en la primera ejecución.")
            print("   Por favor, espera...")

            # Install Playwright browsers
            result = subprocess.run(
                [sys.executable, "-m", "playwright", "install", "chromium"],
                capture_output=True,
                text=True,
                timeout=300  # 5 minutes timeout
            )

            if result.returncode == 0:
                print("✅ Navegadores instalados exitosamente")
                return "Navegadores instalados exitosamente"
            else:
                error_msg = result.stderr or result.stdout
                logger.warning(f"Failed to install browsers: {error_msg}")
                return f"Error al instalar navegadores: {error_msg[:100]}"

        except subprocess.TimeoutExpired:
            return "Timeout al descargar navegadores (toma más de 5 minutos)"
        except Exception as e:
            logger.error(f"Error checking/installing browsers: {e}")
            return f"Error: {e}"

    def _create_first_run_marker(self):
        """Create marker file to indicate first run completed."""
        try:
            marker = self._get_marker_path()
            marker.parent.mkdir(parents=True, exist_ok=True)
            with open(marker, 'w') as f:
                f.write("First run completed\n")
                f.write(f"Config: {self.config._config_path}\n")
                f.write(f"Excel: {self.config.get_excel_file_path()}\n")
                f.write(f"Session: {self.config.get_session_file_path()}\n")
        except Exception as e:
            logger.warning(f"Could not create first run marker: {e}")

    def show_first_run_summary(self):
        """Show summary of first run setup."""
        if not self.config:
            return

        print("\n📋 Resumen de Configuración:")
        print(f"   Archivo de configuración: {self.config._config_path}")
        print(f"   URL de OWA: {self.config.page_url}")
        print(f"   Archivo Excel: {self.config.get_excel_file_path()}")
        print(f"   Archivo de sesión: {self.config.get_session_file_path()}")
        print(f"   Tamaño de lote: {self.config.processing.batch_size}")
        print(f"   Correos por defecto: {len(self.config.default_emails)}")


def check_and_run_first_time_setup(config_path: Optional[str] = None) -> Config:
    """
    Check if first run is needed and run setup if necessary.

    Args:
        config_path: Optional path to configuration file

    Returns:
        Config object (either new or existing)
    """
    first_run = FirstRunManager()

    if first_run.is_first_run():
        print("🔍 Detectada primera ejecución")
        return first_run.run_first_time_setup(config_path)
    else:
        # Normal initialization
        try:
            return Config(config_path)
        except Exception as e:
            logger.warning(f"Config initialization failed: {e}")
            print("⚠️ Error al cargar configuración, intentando configuración inicial...")
            return first_run.run_first_time_setup(config_path)


def show_welcome_message():
    """Show welcome message for GUI applications."""
    welcome_msg = """
🚀 ¡Bienvenido a Verificación de Correos OWA!

Esta es la primera vez que ejecutas la aplicación.
Se te guiará a través de la configuración inicial.

Asegúrate de tener:
• Una cuenta válida en OWA Madrid
• Los correos electrónicos que quieres verificar

El asistente te ayudará a configurar todo lo necesario.
    """

    return welcome_msg.strip()


def get_first_run_instructions() -> str:
    """Get instructions for first-time users."""
    instructions = """
📖 Instrucciones de Primera Ejecución:

1. CONFIGURACIÓN INICIAL:
   • El asistente creará los archivos necesarios automáticamente
   • Revisa la configuración y ajústala según tus necesidades

2. CONFIGURAR SESIÓN:
   • Abre la pestaña "Sesión del Navegador"
   • Haz clic en "Configurar Sesión"
   • Inicia sesión manualmente en OWA
   • Cierra el navegador cuando termines

3. PREPARAR CORREOS:
   • Abre el archivo Excel generado (data/correos.xlsx)
   • Añade los correos que quieres procesar en la columna A
   • La fila 1 debe contener el encabezado "Correo"

4. PROCESAR CORREOS:
   • Ve a la pestaña "Procesamiento"
   • Haz clic en "Iniciar Procesamiento"
   • La aplicación procesará los correos automáticamente

5. VER RESULTADOS:
   • Los resultados se guardan en el mismo archivo Excel
   • Podrás ver los contactos extraídos, teléfonos, etc.

¿Necesitas ayuda? Revisa la documentación o contacta al soporte.
    """

    return instructions.strip()


def install_playwright_browsers() -> bool:
    """
    Install Playwright browsers using the command line interface.
    
    Returns:
        True if installation was successful, False otherwise
    """
    try:
        # Install browsers
        logger.info("Installing Playwright browsers...")
        print("📥 Descargando navegadores de Playwright (primera vez)...")
        print("   Esto puede tomar varios minutos. Por favor, espera...")

        result = subprocess.run(
            [sys.executable, "-m", "playwright", "install", "chromium"],
            capture_output=True,
            text=True,
            timeout=300  # 5 minutes
        )

        if result.returncode == 0:
            logger.info("Playwright browsers installed successfully")
            print("✅ Navegadores instalados exitosamente")
            return True
        else:
            error_msg = result.stderr or result.stdout
            logger.error(f"Failed to install Playwright browsers: {error_msg}")
            print(f"❌ Error al instalar navegadores: {error_msg[:200]}")
            return False

    except subprocess.TimeoutExpired:
        logger.error("Timeout installing Playwright browsers")
        print("❌ Timeout al descargar navegadores (más de 5 minutos)")
        return False
    except Exception as e:
        logger.error(f"Error installing Playwright browsers: {e}")
        print(f"❌ Error al instalar navegadores: {e}")
        return False


def ensure_playwright_browsers_installed() -> bool:
    """
    Ensure Playwright browsers are installed, installing them if necessary.

    This function can be called from anywhere to ensure browsers are available.

    Returns:
        True if browsers are installed or successfully installed, False otherwise
    """
    try:
        from playwright.sync_api import sync_playwright

        # First check if browsers are already installed
        try:
            with sync_playwright() as p:
                browser_path = p.chromium.executable_path
                if browser_path and Path(browser_path).exists():
                    logger.info(f"Playwright browsers already installed at {browser_path}")
                    return True
        except Exception:
            # Browsers not installed, proceed to install
            pass

        return install_playwright_browsers()

    except Exception as e:
        logger.error(f"Error ensuring Playwright browsers: {e}")
        print(f"❌ Error al verificar navegadores: {e}")
        return False