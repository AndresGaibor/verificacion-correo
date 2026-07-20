"""GAL Excel exporter con 2 hojas: Contactos + Compañías.

Sheet1 (Contactos) se actualiza con upsert por persona_id.
Sheet2 (Compañías) lista empresas extraídas para filtrar enrichment.
"""

from pathlib import Path
from typing import List, Dict, Any, Optional
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

CONTACT_FIELDS = ['nombre', 'email', 'empresa', 'telefono', 'departamento', 'oficina', 'direccion', 'persona_id']


def flatten_contact_to_dict(persona: dict) -> dict:
    """Convierte persona del GAL a dict plano."""
    name = persona.get("DisplayName") or ""

    email_obj = persona.get("EmailAddress") or persona.get("EmailAddresses") or []
    email = ""
    if isinstance(email_obj, dict):
        email = email_obj.get("EmailAddress") or ""
    elif isinstance(email_obj, list) and email_obj:
        for e in email_obj:
            val = e.get("Value") or ''
            if val and '@' in val:
                email = val
                break
    elif isinstance(email_obj, str):
        email = email_obj

    phone = ""
    phone_items = persona.get("BusinessPhoneNumbersArray") or []
    if phone_items and isinstance(phone_items[0], dict):
        val = phone_items[0].get("Value") or {}
        if isinstance(val, dict):
            phone = val.get("Number") or val.get("NormalizedNumber") or ""
        elif isinstance(val, str):
            phone = val

    company = persona.get("CompanyName") or ""
    department = persona.get("Department") or ""
    office = persona.get("OfficeLocation") or ""

    address = ""
    addr_items = persona.get("BusinessAddressesArray") or []
    if addr_items and isinstance(addr_items[0], dict):
        val = addr_items[0].get("Value") or {}
        if isinstance(val, dict):
            parts = [val.get("Street") or "", val.get("City") or "", val.get("PostalCode") or "", val.get("State") or ""]
            parts = [p for p in parts if p.strip()]
            address = ", ".join(parts) if parts else ""
        elif isinstance(val, str):
            address = val

    persona_id = ""
    persona_id_obj = persona.get("PersonaId") or {}
    if isinstance(persona_id_obj, dict):
        persona_id = persona_id_obj.get("Id") or ""

    return {
        'nombre': name, 'email': email, 'empresa': company,
        'telefono': phone, 'departamento': department,
        'oficina': office, 'direccion': address, 'persona_id': persona_id,
    }


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 2, 40)


def append_contacts_to_excel(contacts: List[dict], excel_path: Path):
    """Upsert contacts en Sheet1 del Excel por persona_id.

    Sheet2 (Compañías) se recalcula automáticamente.
    Si el archivo no existe, lo crea con las 2 hojas.
    """
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws1 = wb["Contactos"]
    else:
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Contactos"
        ws1.append(CONTACT_FIELDS)
        for cell in ws1[1]:
            cell.font = Font(bold=True)

    existing_rows: Dict[str, int] = {}
    for row_idx in range(2, ws1.max_row + 1):
        pid = ws1.cell(row_idx, 8).value
        if pid:
            existing_rows[pid] = row_idx

    for contact in contacts:
        pid = contact.get('persona_id', '')
        if pid and pid in existing_rows:
            row_idx = existing_rows[pid]
            for col_idx, field in enumerate(CONTACT_FIELDS, 1):
                val = contact.get(field, '')
                if val:
                    ws1.cell(row_idx, col_idx).value = val
        else:
            row = [contact.get(f, '') for f in CONTACT_FIELDS]
            ws1.append(row)
            if pid:
                existing_rows[pid] = ws1.max_row

    if "Compañías" in wb.sheetnames:
        del wb["Compañías"]
    ws2 = wb.create_sheet("Compañías")
    ws2.append(['Compañía', 'Enrich'])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    companies = set()
    for row_idx in range(2, ws1.max_row + 1):
        company = ws1.cell(row_idx, 3).value
        if company:
            companies.add(company)
    for company in sorted(companies):
        ws2.append([company, ''])

    _auto_width(ws1)
    _auto_width(ws2)
    wb.save(excel_path)


def load_gal_from_excel(excel_path: Path) -> List[dict]:
    """Carga contactos desde Sheet1 del Excel."""
    if not excel_path.exists():
        return []
    wb = load_workbook(excel_path)
    ws1 = wb["Contactos"]
    headers = [cell.value for cell in ws1[1]]
    contacts = []
    for row_idx in range(2, ws1.max_row + 1):
        contact = {}
        for col_idx, header in enumerate(headers, 1):
            contact[header] = ws1.cell(row_idx, col_idx).value
        contacts.append(contact)
    return contacts
