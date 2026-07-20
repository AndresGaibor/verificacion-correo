"""
GAL scraper: extracción completa del directorio OWA vía API FindPeople.

Guarda directo a Excel (Sheet1) tras cada página de API, usando upsert por persona_id.
Es reanudable: si el Excel existe, lee max_row para saber desde dónde continuar.

Estrategia:
- Peticiones lentas (delay configurable, default 8s)
- Upsert a Excel tras cada página (no guarda en JSON)
- Detección de sesión expirada (HTTP 307)
"""

import json
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional, Callable
from urllib.request import Request, urlopen
from urllib.error import URLError, HTTPError

from verificacion_correo.utils.logging import get_logger
from verificacion_correo.core.api_extractor import (
    _build_cookie_header,
    _get_canary,
    _build_headers,
    _get_persona,
    validate_session_api,
    SessionExpiredError,
    OWA_BASE,
    SESSION_HEALTH_CHECK_INTERVAL,
    SESSION_ESTIMATED_LIMIT,
)
from verificacion_correo.core.gal_exporter import flatten_contact_to_dict, append_contacts_to_excel

logger = get_logger(__name__)


REQUEST_TIMEOUT = 120
DEFAULT_BATCH_SIZE = 100
DEFAULT_DELAY = 8.0
COMPANIES_CACHE_FILE = "gal_companies.json"


def _build_find_people_payload(
    offset: int,
    max_entries: int,
    address_list_id: str,
    query_string: Optional[str] = None,
) -> Dict[str, Any]:
    """Build the FindPeople request payload for a given offset."""
    body: Dict[str, Any] = {
        "__type": "FindPeopleRequest:#Exchange",
        "IndexedPageItemView": {
            "__type": "IndexedPageView:#Exchange",
            "BasePoint": "Beginning",
            "Offset": offset,
            "MaxEntriesReturned": max_entries,
        },
        "ParentFolderId": {
            "__type": "TargetFolderId:#Exchange",
            "BaseFolderId": {
                "__type": "AddressListId:#Exchange",
                "Id": address_list_id,
            },
        },
        "PersonaShape": {
            "__type": "PersonaResponseShape:#Exchange",
            "BaseShape": "Default",
        },
    }
    if query_string:
        body["QueryString"] = query_string

    return {
        "__type": "FindPeopleJsonRequest:#Exchange",
        "Header": {
            "__type": "JsonRequestHeaders:#Exchange",
            "RequestServerVersion": "Exchange2013",
        },
        "Body": body,
    }


def fetch_company_list(
    session_file: str,
    address_list_id: str = "fed75805-8ba2-4323-9f6d-80be7e3abc6a",
    sample_size: int = 500,
) -> List[str]:
    """Fetch unique company names from GAL via a quick API scan."""
    cookie_str = _build_cookie_header(session_file)
    canary = _get_canary(session_file)
    if not canary:
        raise ValueError("X-OWA-CANARY not found in session file")

    payload = _build_find_people_payload(0, sample_size, address_list_id)
    body_bytes = json.dumps(payload).encode("utf-8")
    req = Request(
        f"{OWA_BASE}/owa/service.svc?action=FindPeople",
        data=body_bytes,
        headers=_build_headers(canary, cookie_str, "FindPeople"),
    )

    try:
        resp = urlopen(req, timeout=REQUEST_TIMEOUT)
        data = json.loads(resp.read().decode("utf-8"))
        body = data.get("Body", {})
        people = body.get("People") or body.get("ResultSet") or []

        companies = set()
        for person in people:
            company = (person.get("CompanyName") or "").strip()
            if company:
                companies.add(company)

        return sorted(companies)

    except HTTPError as e:
        if e.code == 307:
            raise SessionExpiredError("Session expired during company list fetch")
        raise
    except Exception as e:
        logger.error(f"Error fetching company list: {e}")
        return []


def save_companies_cache(companies: List[str], output_dir: Path):
    """Save company list to cache file."""
    cache_path = output_dir / COMPANIES_CACHE_FILE
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(companies, f, ensure_ascii=False, indent=2)
    logger.info(f"Companies cache saved: {cache_path} ({len(companies)} companies)")


def load_companies_cache(output_dir: Path) -> Optional[List[str]]:
    """Load company list from cache file."""
    cache_path = output_dir / COMPANIES_CACHE_FILE
    if cache_path.exists():
        with open(cache_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return None


def _enrich_persona(persona: dict, cookie_str: str, canary: str) -> dict:
    """Enrich a single persona with GetPersona details (phone, dept, office, address)."""
    persona_id = ""
    persona_id_obj = persona.get("PersonaId") or {}
    if isinstance(persona_id_obj, dict):
        persona_id = persona_id_obj.get("Id") or ""

    if not persona_id:
        return persona

    try:
        enriched = _get_persona(persona_id, cookie_str, canary)
        if enriched:
            persona["BusinessPhoneNumbersArray"] = [{"Value": {"Number": enriched.phone}}] if enriched.phone else []
            persona["Department"] = enriched.department or ""
            persona["OfficeLocation"] = enriched.office_location or ""
            persona["BusinessAddressesArray"] = [{"Value": enriched.address}] if enriched.address else []
            if enriched.name and not persona.get("DisplayName"):
                persona["DisplayName"] = enriched.name
    except Exception as e:
        logger.debug(f"GetPersona enrichment failed for {persona_id}: {e}")

    return persona


def scrape_gal(
    session_file: str,
    excel_path: str,
    max_contacts: int = 0,
    batch_size: int = DEFAULT_BATCH_SIZE,
    request_delay: float = DEFAULT_DELAY,
    address_list_id: str = "fed75805-8ba2-4323-9f6d-80be7e3abc6a",
    force_restart: bool = False,
    progress_callback: Optional[Callable[[int, int], None]] = None,
    session_health_callback: Optional[Callable[[dict], None]] = None,
    stop_flag: Optional[Dict[str, bool]] = None,
    company_filter: Optional[List[str]] = None,
    enrich_contacts: bool = False,
) -> Dict[str, Any]:
    """Extrae directorio GAL directo a Excel (Sheet1) con upsert por persona_id.

    Args:
        session_file: Path a archivo de sesión con cookies.
        excel_path: Path al Excel de salida (Sheet1=Contactos, Sheet2=Compañías).
        max_contacts: Máx contactos a extraer (0=ilimitado).
        batch_size: Entradas por página.
        request_delay: Segundos entre peticiones.
        address_list_id: GAL address list GUID.
        force_restart: Si True, sobrescribe Excel existente.
        progress_callback: fn(count, total) para actualizar UI.
        session_health_callback: fn(health_info) para salud de sesión.
        stop_flag: dict{"stop": bool} para señalar parada.
        company_filter: Lista de compañías para filtrado client-side.
        enrich_contacts: Si True, fetch GetPersona para cada contacto.
    """
    if not session_file or not Path(session_file).exists():
        raise FileNotFoundError(f"Session file not found: {session_file}")
    if not excel_path:
        raise ValueError("excel_path cannot be empty")
    if batch_size <= 0:
        raise ValueError(f"batch_size must be > 0, got {batch_size}")

    start = time.time()
    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    cookie_str = _build_cookie_header(session_file)
    canary = _get_canary(session_file)
    if not canary:
        raise ValueError("X-OWA-CANARY not found in session file")

    if session_health_callback:
        initial_health = validate_session_api(session_file)
        initial_health['calls_used'] = 0
        initial_health['estimated_limit'] = SESSION_ESTIMATED_LIMIT
        session_health_callback(initial_health)
        if not initial_health['valid']:
            raise SessionExpiredError(f"Session validation failed: {initial_health['message']}")

    resume_offset = 0
    if not force_restart and excel_path.exists():
        from openpyxl import load_workbook
        try:
            wb = load_workbook(excel_path)
            ws1 = wb["Contactos"]
            resume_offset = ws1.max_row - 1
            logger.info(f"Resuming from row {resume_offset + 1} (max_row={ws1.max_row})")
        except Exception:
            resume_offset = 0

    session_expired = False
    total_fetched = resume_offset
    total_scanned = resume_offset
    consecutive_failures = 0
    max_consecutive_failures = 5
    api_calls = 0

    def _check_session_health():
        nonlocal session_expired, api_calls
        if session_health_callback and api_calls % SESSION_HEALTH_CHECK_INTERVAL == 0:
            health = validate_session_api(session_file)
            health['calls_used'] = api_calls
            health['estimated_limit'] = SESSION_ESTIMATED_LIMIT
            session_health_callback(health)
            if not health['valid']:
                logger.warning(f"Session health check failed: {health['message']}")
                session_expired = True
                return True
        return False

    def _fetch_page(offset: int, current_batch: int, query_string: Optional[str] = None) -> List[dict]:
        payload = _build_find_people_payload(offset, current_batch, address_list_id, query_string)
        body_bytes = json.dumps(payload).encode("utf-8")
        req = Request(
            f"{OWA_BASE}/owa/service.svc?action=FindPeople",
            data=body_bytes,
            headers=_build_headers(canary, cookie_str, "FindPeople"),
        )
        resp = urlopen(req, timeout=REQUEST_TIMEOUT)
        data = json.loads(resp.read().decode("utf-8"))
        body = data.get("Body", {})
        return body.get("People") or body.get("ResultSet") or []

    def _enrich_batch(people_list: List[dict]):
        if not enrich_contacts or not people_list:
            return
        for i, person in enumerate(people_list):
            people_list[i] = _enrich_persona(person, cookie_str, canary)
            if i < len(people_list) - 1:
                time.sleep(0.5)

    company_filter_enabled = bool(company_filter)
    company_set = set(company_filter) if company_filter else set()

    if company_filter_enabled:
        effective_batch = max(batch_size, 1000)
        logger.info(
            f"Filtered mode: fetching all GAL with batch={effective_batch}, "
            f"filtering client-side by {company_filter}"
        )
        scanned_offset = resume_offset
        total_scanned = resume_offset

        while True:
            if session_expired or (stop_flag and stop_flag.get("stop")):
                break
            if max_contacts > 0 and total_fetched >= max_contacts:
                break

            current_batch = effective_batch
            remaining = max_contacts - total_fetched if max_contacts > 0 else 0
            if remaining > 0:
                current_batch = min(effective_batch, remaining)

            try:
                people = _fetch_page(scanned_offset, current_batch)
                if not people:
                    logger.info("No more results from GAL")
                    break

                consecutive_failures = 0
                api_calls += 1
                total_scanned += len(people)

                _enrich_batch(people)

                filtered = [p for p in people if p.get("CompanyName") in company_set]
                if filtered:
                    flattened = [flatten_contact_to_dict(p) for p in filtered]
                    append_contacts_to_excel(flattened, excel_path)
                    total_fetched += len(filtered)

                scanned_offset += len(people)

                if _check_session_health():
                    break

                if progress_callback:
                    progress_callback(total_fetched, max_contacts)

                if stop_flag and stop_flag.get("stop"):
                    break

                logger.info(
                    f"Scanned {total_scanned} total, {total_fetched} matched from {company_filter}"
                )
                time.sleep(request_delay)

            except HTTPError as e:
                if e.code == 307:
                    logger.warning(f"Session expired at offset {scanned_offset}")
                    session_expired = True
                    break
                consecutive_failures += 1
                logger.error(f"HTTP {e.code} at offset {scanned_offset}: {e.reason} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

            except URLError as e:
                consecutive_failures += 1
                logger.error(f"Connection error at offset {scanned_offset}: {e.reason} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

            except Exception as e:
                consecutive_failures += 1
                logger.error(f"Unexpected error at offset {scanned_offset}: {e} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

    else:
        offset = resume_offset

        while True:
            if stop_flag and stop_flag.get("stop"):
                break
            if max_contacts > 0 and total_fetched >= max_contacts:
                break

            current_batch = batch_size
            if max_contacts > 0:
                remaining = max_contacts - total_fetched
                if remaining <= 0:
                    break
                current_batch = min(batch_size, remaining)

            logger.info(f"Fetching offset {offset} (batch size {current_batch})...")

            try:
                people = _fetch_page(offset, current_batch)
                if not people:
                    logger.info("No more results — GAL fully fetched")
                    break

                consecutive_failures = 0
                total_scanned += len(people)

                _enrich_batch(people)

                flattened = [flatten_contact_to_dict(p) for p in people]
                append_contacts_to_excel(flattened, excel_path)

                total_fetched += len(people)
                offset += len(people)
                api_calls += 1

                if _check_session_health():
                    break

                if progress_callback:
                    progress_callback(total_fetched, max_contacts)

                if stop_flag and stop_flag.get("stop"):
                    break

                time.sleep(request_delay)

            except HTTPError as e:
                if e.code == 307:
                    logger.warning(f"Session expired at offset {offset}")
                    session_expired = True
                    break
                consecutive_failures += 1
                logger.error(f"HTTP {e.code} at offset {offset}: {e.reason} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

            except URLError as e:
                consecutive_failures += 1
                logger.error(f"Connection error at offset {offset}: {e.reason} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

            except Exception as e:
                consecutive_failures += 1
                logger.error(f"Unexpected error at offset {offset}: {e} (fail {consecutive_failures}/{max_consecutive_failures})")
                if consecutive_failures >= max_consecutive_failures:
                    logger.error("Too many consecutive failures, aborting")
                    break
                time.sleep(request_delay * 2)

    if session_health_callback:
        final_health = validate_session_api(session_file) if not session_expired else {'valid': False, 'health': 'expired', 'message': 'Sesión expirada'}
        final_health['calls_used'] = api_calls
        final_health['estimated_limit'] = SESSION_ESTIMATED_LIMIT
        session_health_callback(final_health)

    duration = time.time() - start

    stats = {
        "total": total_fetched,
        "total_scanned": total_scanned,
        "offset_end": total_scanned,
        "expired": session_expired,
        "stopped": stop_flag and stop_flag.get("stop"),
        "files": {"excel": str(excel_path)},
        "duration": duration,
        "api_calls": api_calls,
    }

    if company_filter:
        stats["filtered_companies"] = company_filter

    if session_expired:
        logger.warning(f"Scraper stopped early: session expired ({total_fetched} contacts)")
    elif stop_flag and stop_flag.get("stop"):
        logger.info(f"Scraper stopped by user: {total_fetched} contacts")
    else:
        logger.info(f"GAL scraping complete: {total_fetched} contacts in {duration:.1f}s")

    return stats
