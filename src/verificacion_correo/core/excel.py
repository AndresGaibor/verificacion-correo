"""
Excel file operations for verificacion-correo.

This module provides functionality to read email addresses from Excel files
and write results back to the same files, with proper error handling and
structured data management.
"""

import os
import sys
import shutil
from enum import Enum
from pathlib import Path
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

from verificacion_correo.utils.logging import get_logger


logger = get_logger(__name__)


class ProcessingStatus(Enum):
    """Processing status enumeration."""
    PENDING = ""
    SUCCESS = "OK"
    NOT_FOUND = "NO EXISTE"
    ERROR = "ERROR"


@dataclass
class ExcelColumn:
    """Excel column definition."""
    name: str
    letter: str
    index: int
    description: str


class ExcelColumns:
    """Excel column definitions for contact data."""

    EMAIL = ExcelColumn("Correo", "A", 1, "Email address to process")
    STATUS = ExcelColumn("Status", "B", 2, "Processing status")
    NAME = ExcelColumn("Nombre", "C", 3, "Contact name")
    EMAIL_PERSONAL = ExcelColumn("Email Personal", "D", 4, "Personal email")
    PHONE = ExcelColumn("Teléfono", "E", 5, "Phone number")
    SIP = ExcelColumn("SIP", "F", 6, "SIP address")
    ADDRESS = ExcelColumn("Dirección", "G", 7, "Physical address")
    DEPARTMENT = ExcelColumn("Departamento", "H", 8, "Department/Office")
    COMPANY = ExcelColumn("Compañía", "I", 9, "Company/Organization")
    OFFICE_LOCATION = ExcelColumn("Oficina", "J", 10, "Office location")

    @classmethod
    def get_all_columns(cls) -> List[ExcelColumn]:
        """Get all column definitions."""
        return [
            cls.EMAIL, cls.STATUS, cls.NAME, cls.EMAIL_PERSONAL,
            cls.PHONE, cls.SIP, cls.ADDRESS, cls.DEPARTMENT,
            cls.COMPANY, cls.OFFICE_LOCATION
        ]

    @classmethod
    def get_headers(cls) -> List[str]:
        """Get list of header names."""
        return [col.name for col in cls.get_all_columns()]

    @classmethod
    def get_column_by_name(cls, name: str) -> Optional[ExcelColumn]:
        """Get column definition by name."""
        for col in cls.get_all_columns():
            if col.name == name:
                return col
        return None


@dataclass
class EmailRecord:
    """Email record with metadata."""
    email: str
    row: int
    status: ProcessingStatus = ProcessingStatus.PENDING
    data: Optional[Dict[str, Any]] = None


@dataclass
class ExcelSummary:
    """Summary of Excel file contents."""
    total_emails: int
    pending_count: int
    processed_count: int
    batches: List[List[EmailRecord]]


class ExcelReader:
    """
    Excel file reader with batch processing support.

    Reads email addresses and processing status from Excel files,
    organizing them into batches for efficient processing.
    """

    def __init__(self, file_path: str, start_row: int = 2, email_column: int = 1):
        """
        Initialize Excel reader.

        Args:
            file_path: Path to Excel file
            start_row: Row where data starts (default: 2)
            email_column: Column index for emails (default: 1 for column A)
        """
        self.file_path = Path(file_path)
        self.start_row = start_row
        self.email_column = email_column
        self.status_column = ExcelColumns.STATUS.index

        # Ensure Excel file exists (create from template if needed)
        self._ensure_excel_file_exists()

    def _ensure_excel_file_exists(self):
        """Ensure Excel file exists, create from template if needed."""
        if not self.file_path.exists():
            logger.info(f"Excel file not found: {self.file_path}")

            # Try to create from template first
            if self._create_from_template():
                return

            # Create empty file with structure if template not available
            self._create_empty_excel()

    def _create_from_template(self) -> bool:
        """Try to create Excel file from bundled template."""
        try:
            # Check if we're running in PyInstaller bundle
            if getattr(sys, 'frozen', False) and hasattr(sys, '_MEIPASS'):
                bundle_dir = Path(sys._MEIPASS)
                template_path = bundle_dir / "data" / "correos_template.xlsx"

                if template_path.exists():
                    logger.info(f"Creating Excel file from template: {template_path}")
                    # Ensure parent directory exists
                    self.file_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy(template_path, self.file_path)
                    logger.info(f"Excel file created from template: {self.file_path}")
                    return True

            # Try to find template in development
            template_paths = [
                Path("data/correos_template.xlsx"),
                Path("data/correos.xlsx"),  # fallback to existing file
            ]

            for template_path in template_paths:
                if template_path.exists():
                    logger.info(f"Creating Excel file from template: {template_path}")
                    self.file_path.parent.mkdir(parents=True, exist_ok=True)
                    shutil.copy(template_path, self.file_path)
                    logger.info(f"Excel file created from template: {self.file_path}")
                    return True

        except Exception as e:
            logger.warning(f"Failed to create Excel from template: {e}")

        return False

    def _create_empty_excel(self):
        """Create an empty Excel file with basic structure."""
        logger.info(f"Creating empty Excel file: {self.file_path}")

        try:
            # Ensure parent directory exists
            self.file_path.parent.mkdir(parents=True, exist_ok=True)

            wb = Workbook()
            ws = wb.active
            ws.title = "Contactos"

            # Add headers
            headers = ExcelColumns.get_headers()
            for col_idx, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col_idx, value=header)

            # Add sample data
            sample_emails = [
                'ASP164@MADRID.ORG',
                'AGM564@MADRID.ORG',
                'USR789@MADRID.ORG'
            ]

            for row_idx, email in enumerate(sample_emails, start=2):
                ws.cell(row=row_idx, column=1, value=email)  # Email column
                ws.cell(row=row_idx, column=2, value="")     # Status column (empty = pending)

            # Auto-adjust column widths
            for col in range(1, len(headers) + 1):
                column_letter = get_column_letter(col)
                ws.column_dimensions[column_letter].width = 20

            wb.save(self.file_path)
            logger.info(f"Empty Excel file created with sample data: {self.file_path}")

        except Exception as e:
            logger.error(f"Failed to create empty Excel file: {e}")
            # Create parent directory at minimum
            self.file_path.parent.mkdir(parents=True, exist_ok=True)

    def read_all_emails(self) -> List[str]:
        """
        Read all email addresses from the file.

        Returns:
            List of email addresses found in the file.
        """
        if not self.file_path.exists():
            logger.warning(f"Excel file not found: {self.file_path}")
            return []

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            emails = []
            for row in range(self.start_row, ws.max_row + 1):
                cell = ws.cell(row=row, column=self.email_column)
                value = cell.value

                if value:
                    email = str(value).strip()
                    if email and '@' in email:
                        emails.append(email)

            wb.close()
            logger.info(f"Read {len(emails)} emails from {self.file_path}")
            return emails

        except Exception as e:
            logger.error(f"Error reading Excel file {self.file_path}: {e}")
            return []

    def read_pending_emails(self, batch_size: int = 10) -> ExcelSummary:
        """
        Read pending emails and organize into batches.

        Args:
            batch_size: Number of emails per batch

        Returns:
            ExcelSummary with batches and statistics.
        """
        if not self.file_path.exists():
            logger.error(f"Excel file not found: {self.file_path}")
            return ExcelSummary(0, 0, 0, [])

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            pending_records = []
            total_emails = 0
            processed_count = 0

            for row in range(self.start_row, ws.max_row + 1):
                email_cell = ws.cell(row=row, column=self.email_column)
                status_cell = ws.cell(row=row, column=self.status_column)

                email = email_cell.value
                status = status_cell.value

                if email:
                    email = str(email).strip()
                    if email and '@' in email:
                        total_emails += 1

                        # Check if processing is needed
                        if not status or str(status).strip() == "":
                            record = EmailRecord(
                                email=email,
                                row=row,
                                status=ProcessingStatus.PENDING
                            )
                            pending_records.append(record)
                        else:
                            processed_count += 1

            wb.close()

            # Create batches
            batches = []
            for i in range(0, len(pending_records), batch_size):
                batch = pending_records[i:i + batch_size]
                batches.append(batch)

            summary = ExcelSummary(
                total_emails=total_emails,
                pending_count=len(pending_records),
                processed_count=processed_count,
                batches=batches
            )

            logger.info(
                f"Found {summary.pending_count} pending emails in {summary.total_emails} total, "
                f"organized into {len(batches)} batches"
            )

            return summary

        except Exception as e:
            logger.error(f"Error reading pending emails from {self.file_path}: {e}")
            return ExcelSummary(0, 0, 0, [])


class ExcelWriter:
    """
    Excel file writer for recording processing results.

    Writes processing status and extracted contact data back to Excel files.
    """

    def __init__(self, file_path: str):
        """
        Initialize Excel writer.

        Args:
            file_path: Path to Excel file
        """
        self.file_path = Path(file_path)
        self.columns = ExcelColumns()

    def ensure_file_structure(self):
        """
        Ensure the Excel file exists and has proper headers.

        Creates the file if it doesn't exist and ensures all headers are present.
        """
        # Create directory if it doesn't exist
        self.file_path.parent.mkdir(parents=True, exist_ok=True)

        if not self.file_path.exists():
            self._create_new_file()
        else:
            self._verify_headers()

    def _create_new_file(self):
        """Create a new Excel file with headers."""
        logger.info(f"Creating new Excel file: {self.file_path}")

        wb = Workbook()
        ws = wb.active
        ws.title = "Contactos"

        # Write headers
        headers = self.columns.get_headers()
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        wb.save(self.file_path)

    def _verify_headers(self):
        """Verify and fix headers if necessary."""
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            headers_updated = False
            headers = self.columns.get_headers()

            for col_idx, expected_header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx)
                if cell.value != expected_header:
                    logger.warning(f"Updating header in column {col_idx}: '{cell.value}' -> '{expected_header}'")
                    cell.value = expected_header
                    headers_updated = True

            if headers_updated:
                wb.save(self.file_path)
                logger.info(f"Updated headers in {self.file_path}")

            wb.close()

        except Exception as e:
            logger.error(f"Error verifying headers in {self.file_path}: {e}")

    def write_result(self, record: EmailRecord):
        """
        Write processing result for a single email record.

        Args:
            record: EmailRecord with processing results
        """
        self.ensure_file_structure()

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            # Write status
            status_cell = ws.cell(row=record.row, column=self.columns.STATUS.index)
            status_cell.value = record.status.value

            # Write data if processing was successful
            if record.status == ProcessingStatus.SUCCESS and record.data:
                self._write_contact_data(ws, record)
            elif record.status in [ProcessingStatus.ERROR, ProcessingStatus.NOT_FOUND]:
                self._clear_contact_data(ws, record)

            wb.save(self.file_path)
            wb.close()

            logger.debug(f"Wrote result for {record.email}: {record.status.value}")

        except Exception as e:
            logger.error(f"Error writing result for {record.email}: {e}")

    def _write_contact_data(self, ws, record: EmailRecord):
        """Write contact data to worksheet."""
        data_mapping = {
            'name': self.columns.NAME.index,
            'email': self.columns.EMAIL_PERSONAL.index,
            'phone': self.columns.PHONE.index,
            'sip': self.columns.SIP.index,
            'address': self.columns.ADDRESS.index,
            'department': self.columns.DEPARTMENT.index,
            'company': self.columns.COMPANY.index,
            'office_location': self.columns.OFFICE_LOCATION.index,
        }

        for field, column in data_mapping.items():
            if field in record.data and record.data[field]:
                ws.cell(row=record.row, column=column, value=record.data[field])

    def _clear_contact_data(self, ws, record: EmailRecord):
        """Clear contact data columns for error/not found records."""
        start_col = self.columns.NAME.index
        end_col = self.columns.OFFICE_LOCATION.index

        for col in range(start_col, end_col + 1):
            ws.cell(row=record.row, column=col, value=None)

    def write_batch_results(self, records: List[EmailRecord]):
        """
        Write results for a batch of email records.

        Args:
            records: List of EmailRecord objects
        """
        self.ensure_file_structure()

        try:
            wb = load_workbook(self.file_path)
            ws = wb.active

            for record in records:
                # Write status
                status_cell = ws.cell(row=record.row, column=self.columns.STATUS.index)
                status_cell.value = record.status.value

                # Write data if processing was successful
                if record.status == ProcessingStatus.SUCCESS and record.data:
                    self._write_contact_data(ws, record)
                elif record.status in [ProcessingStatus.ERROR, ProcessingStatus.NOT_FOUND]:
                    self._clear_contact_data(ws, record)

            wb.save(self.file_path)
            wb.close()

            successful = sum(1 for r in records if r.status == ProcessingStatus.SUCCESS)
            logger.info(f"Wrote batch results: {successful}/{len(records)} successful")

        except Exception as e:
            logger.error(f"Error writing batch results: {e}")

    def get_status(self, row: int) -> Optional[str]:
        """
        Get the processing status for a specific row.

        Args:
            row: Row number (1-based)

        Returns:
            Status string or None if not found
        """
        if not self.file_path.exists():
            return None

        try:
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active

            status = ws.cell(row=row, column=self.columns.STATUS.index).value

            wb.close()
            return status

        except Exception as e:
            logger.error(f"Error reading status from row {row}: {e}")
            return None