"""GAL enrichment selectivo por compañía usando GetPersona API."""

from pathlib import Path
from typing import List, Dict, Any, Set, Optional, Callable
import json
import time
from openpyxl import load_workbook
from urllib.request import Request, urlopen
from urllib.error import HTTPError

from .api_extractor import _build_cookie_header, _get_canary, _build_headers, OWA_BASE


def _call_get_persona(persona_id: str, cookie_str: str, canary: str) -> Optional[dict]:
    """Llama GetPersona API y retorna dict con datos enriquecidos."""
    payload = {
        "__type": "GetPersonaJsonRequest:#Exchange",
        "Header": {
            "__type": "JsonRequestHeaders:#Exchange",
            "RequestServerVersion": "Exchange2013",
        },
        "Body": {
            "__type": "GetPersonaRequest:#Exchange",
            "PersonaId": {
                "__type": "ItemId:#Exchange",
                "Id": persona_id,
            },
            "PersonaShape": {
                "__type": "PersonaResponseShape:#Exchange",
                "BaseShape": "Default",
            },
        },
    }

    body_bytes = json.dumps(payload).encode("utf-8")
    req = Request(
        f"{OWA_BASE}/owa/service.svc?action=GetPersona",
        data=body_bytes,
        headers=_build_headers(canary, cookie_str, "GetPersona"),
    )

    try:
        resp = urlopen(req, timeout=60)
        data = json.loads(resp.read().decode("utf-8"))
        persona = data.get("Body", {}).get("Persona")
        if not persona:
            return None

        phone = ""
        phone_items = persona.get("BusinessPhoneNumbersArray") or []
        if phone_items and isinstance(phone_items[0], dict):
            val = phone_items[0].get("Value") or {}
            if isinstance(val, dict):
                phone = val.get("Number") or val.get("NormalizedNumber") or ""

        address = ""
        addr_items = persona.get("BusinessAddressesArray") or []
        if addr_items and isinstance(addr_items[0], dict):
            val = addr_items[0].get("Value") or {}
            if isinstance(val, dict):
                parts = [
                    val.get("Street") or "",
                    val.get("City") or "",
                    val.get("PostalCode") or "",
                    val.get("State") or "",
                ]
                parts = [p for p in parts if p.strip()]
                address = ", ".join(parts)

        return {
            "telefono": phone,
            "departamento": persona.get("Department") or "",
            "oficina": "",
            "direccion": address,
        }
    except HTTPError as e:
        if e.code == 307:
            raise Exception("Session expired during GetPersona")
        return None
    except Exception:
        return None


class EnrichProgress:
    """Trackea progreso de enrichment para reanudación."""

    def __init__(self, path: Path):
        self.path = path
        self.data: Dict[str, Any] = {
            'companies_done': [],
            'contacts_enriched': 0,
            'offset': 0,
        }

    def save(self):
        with open(self.path, 'w', encoding='utf-8') as f:
            json.dump(self.data, f, indent=2, ensure_ascii=False)

    def load(self) -> bool:
        if self.path.exists():
            with open(self.path, 'r', encoding='utf-8') as f:
                self.data = json.load(f)
            return True
        return False


def find_contact_in_cache(email: str, empresa: str, cache: List[dict]) -> Optional[dict]:
    """Busca contacto exacto en cache por email + empresa."""
    for c in cache:
        if c.get('email', '').lower() == email.lower() and c.get('empresa', '') == empresa:
            return c
    return None


def merge_enrichment(existing: dict, enrichment: dict) -> dict:
    """Mezcla enrichment en existing - solo llena vacíos."""
    result = existing.copy()
    for key in ['telefono', 'departamento', 'oficina', 'direccion']:
        if not result.get(key) and enrichment.get(key):
            result[key] = enrichment.get(key)
    return result


def enrich_excel_by_companies(
    excel_path: Path,
    companies: List[str],
    cache: List[dict],
    session_file: str = "state.json",
    progress_path: Optional[Path] = None,
    progress_callback: Optional[Callable[[int, int], None]] = None,
) -> Dict[str, Any]:
    """Enriquece Excel selectivamente por lista de compañías usando GetPersona API.

    1. Lee Sheet2 -> filtra por companies con X
    2. Para cada compañía:
       a. Busca en Sheet1 filas donde Empresa == compañía
       b. Para cada match -> busca persona_id en cache (usa cache raw si tiene PersonaId)
       c. Si tiene persona_id -> llama GetPersona API -> llena vacíos
    3. Guarda Excel actualizado
    """
    companies_set = set(companies)
    progress = EnrichProgress(progress_path) if progress_path else None
    if progress and progress.load():
        companies_set -= set(progress.data.get('companies_done', []))

    try:
        cookie_str = _build_cookie_header(session_file)
        canary = _get_canary(session_file)
    except Exception:
        return {
            'companies_done': 0,
            'contacts_enriched': 0,
            'error': 'No se pudo cargar la sesión'
        }

    wb = load_workbook(excel_path)
    ws1 = wb["Contactos"]
    ws2 = wb["Compañías"]

    cache_index: Dict[tuple, dict] = {}
    for c in cache:
        email_obj = c.get('EmailAddress') or c.get('email', {})
        if isinstance(email_obj, dict):
            email = email_obj.get('EmailAddress', '').lower()
        elif isinstance(email_obj, str):
            email = email_obj.lower()
        else:
            email = ''

        empresa = c.get('CompanyName') or c.get('empresa', '')
        key = (email, empresa)
        cache_index[key] = c

    headers = [cell.value for cell in ws1[1]]
    telefono_idx = headers.index('telefono') + 1
    depto_idx = headers.index('departamento') + 1
    oficina_idx = headers.index('oficina') + 1
    direccion_idx = headers.index('direccion') + 1
    empresa_idx = headers.index('empresa') + 1
    email_idx = headers.index('email') + 1

    total_enriched = 0
    companies_done: List[str] = []
    errors = []

    for row_idx in range(2, ws2.max_row + 1):
        company = ws2.cell(row_idx, 1).value
        enrich_mark = ws2.cell(row_idx, 2).value

        enrich_str = str(enrich_mark).strip().upper() if enrich_mark else ''
        if not company or enrich_str != 'X':
            continue
        if company not in companies_set:
            continue

        for data_row_idx in range(2, ws1.max_row + 1):
            empresa_cell = ws1.cell(data_row_idx, empresa_idx).value
            if empresa_cell != company:
                continue

            email_cell = ws1.cell(data_row_idx, email_idx).value
            cache_key = (str(email_cell).lower() if email_cell else '', company)
            cached = cache_index.get(cache_key)

            if not cached:
                continue

            persona_id = cached.get('persona_id', '')
            if not persona_id:
                persona_id_obj = cached.get('PersonaId') or {}
                if isinstance(persona_id_obj, dict):
                    persona_id = persona_id_obj.get('Id', '')

            if not persona_id:
                continue

            try:
                enriched = _call_get_persona(persona_id, cookie_str, canary)
            except Exception as e:
                errors.append(str(e))
                continue

            if not enriched:
                continue

            telefono = ws1.cell(data_row_idx, telefono_idx).value
            if not telefono and enriched.get('telefono'):
                ws1.cell(data_row_idx, telefono_idx).value = enriched['telefono']

            depto = ws1.cell(data_row_idx, depto_idx).value
            if not depto and enriched.get('departamento'):
                ws1.cell(data_row_idx, depto_idx).value = enriched['departamento']

            oficina = ws1.cell(data_row_idx, oficina_idx).value
            if not oficina and enriched.get('oficina'):
                ws1.cell(data_row_idx, oficina_idx).value = enriched['oficina']

            direccion = ws1.cell(data_row_idx, direccion_idx).value
            if not direccion and enriched.get('direccion'):
                ws1.cell(data_row_idx, direccion_idx).value = enriched['direccion']

            total_enriched += 1

            if progress_callback:
                progress_callback(total_enriched, 0)

            time.sleep(0.5)

        companies_done.append(company)

        if progress:
            progress.data['companies_done'] = companies_done
            progress.data['contacts_enriched'] = total_enriched
            progress.save()

    wb.save(excel_path)

    result = {
        'companies_done': len(companies_done),
        'contacts_enriched': total_enriched,
    }
    if errors:
        result['errors'] = errors[:5]

    return result


def get_companies_to_enrich_from_excel(excel_path: Path) -> List[str]:
    """Lee Sheet2 y retorna lista de compañías marcadas con X."""
    wb = load_workbook(excel_path)
    ws2 = wb["Compañías"]
    companies = []
    for row_idx in range(2, ws2.max_row + 1):
        company = ws2.cell(row_idx, 1).value
        enrich_mark = ws2.cell(row_idx, 2).value
        enrich_str = str(enrich_mark).strip().upper() if enrich_mark else ''
        if company and enrich_str == 'X':
            companies.append(company)
    return companies
