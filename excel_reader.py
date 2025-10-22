"""
Módulo para leer correos electrónicos desde archivos Excel.
"""

import os
from openpyxl import load_workbook
from config import DEFAULT_EMAILS, EXCEL_CONFIG, PROCESSING_CONFIG
from excel_writer import crear_headers_si_no_existen


def leer_correos_excel(archivo_path=None):
    """
    Lee correos electrónicos desde un archivo Excel.
    Lee la primera columna (A) desde la fila 2 en adelante.

    Args:
        archivo_path: Ruta al archivo Excel (por defecto usa EXCEL_CONFIG['default_file'])

    Returns:
        Lista de correos electrónicos encontrados
    """
    if archivo_path is None:
        archivo_path = EXCEL_CONFIG['default_file']

    correos = []

    # Verificar que el archivo existe
    if not os.path.exists(archivo_path):
        print(f"ADVERTENCIA: No se encontró el archivo {archivo_path}")
        print("Usando correos por defecto.")
        return DEFAULT_EMAILS

    try:
        # Cargar el archivo Excel
        wb = load_workbook(archivo_path, read_only=True, data_only=True)
        ws = wb.active  # Usar la hoja activa

        # Leer la primera columna desde la fila configurada
        start_row = EXCEL_CONFIG['start_row']
        email_column = EXCEL_CONFIG['email_column']

        for row in range(start_row, ws.max_row + 1):
            celda = ws.cell(row=row, column=email_column)
            valor = celda.value

            # Verificar que la celda no esté vacía
            if valor:
                # Convertir a string y limpiar espacios
                correo = str(valor).strip()
                if correo:  # Solo agregar si no está vacío después de strip
                    correos.append(correo)

        wb.close()

        if not correos:
            print(f"ADVERTENCIA: No se encontraron correos en {archivo_path}")
            print("Usando correos por defecto.")
            return DEFAULT_EMAILS

        print(f"Se cargaron {len(correos)} correos desde {archivo_path}")
        return correos

    except Exception as e:
        print(f"ERROR al leer el archivo Excel: {e}")
        print("Usando correos por defecto.")
        return DEFAULT_EMAILS


def leer_correos_pendientes(archivo_path=None):
    """
    Lee correos electrónicos que necesitan ser consultados (columna Status vacía).
    Retorna los emails organizados en lotes para procesamiento eficiente.

    Args:
        archivo_path: Ruta al archivo Excel (por defecto usa EXCEL_CONFIG['default_file'])

    Returns:
        Diccionario con:
        - 'lotes': Lista de lotes, cada lote es una lista de dicts {'email': str, 'fila': int}
        - 'total_pendientes': Número total de emails pendientes
        - 'total_procesados': Número de emails ya procesados
        - 'total_emails': Número total de emails en el archivo
    """
    if archivo_path is None:
        archivo_path = EXCEL_CONFIG['default_file']

    # Asegurar que el archivo tenga headers
    crear_headers_si_no_existen(archivo_path)

    # Verificar que el archivo existe
    if not os.path.exists(archivo_path):
        print(f"ERROR: No se encontró el archivo {archivo_path}")
        return {
            'lotes': [],
            'total_pendientes': 0,
            'total_procesados': 0,
            'total_emails': 0
        }

    try:
        # Cargar el archivo Excel
        wb = load_workbook(archivo_path, read_only=True, data_only=True)
        ws = wb.active

        emails_pendientes = []
        total_procesados = 0
        total_emails = 0

        start_row = EXCEL_CONFIG['start_row']
        email_column = EXCEL_CONFIG['email_column']
        status_column = 2  # Columna B (Status)

        # Leer todas las filas
        for row in range(start_row, ws.max_row + 1):
            email_celda = ws.cell(row=row, column=email_column)
            status_celda = ws.cell(row=row, column=status_column)

            email = email_celda.value
            status = status_celda.value

            # Si hay email en la celda
            if email:
                email = str(email).strip()
                if email:
                    total_emails += 1

                    # Si el status está vacío, necesita ser consultado
                    if not status or str(status).strip() == '':
                        emails_pendientes.append({
                            'email': email,
                            'fila': row
                        })
                    else:
                        total_procesados += 1

        wb.close()

        # Dividir en lotes
        batch_size = PROCESSING_CONFIG['batch_size']
        lotes = []

        for i in range(0, len(emails_pendientes), batch_size):
            lote = emails_pendientes[i:i + batch_size]
            lotes.append(lote)

        return {
            'lotes': lotes,
            'total_pendientes': len(emails_pendientes),
            'total_procesados': total_procesados,
            'total_emails': total_emails
        }

    except Exception as e:
        print(f"ERROR al leer el archivo Excel: {e}")
        return {
            'lotes': [],
            'total_pendientes': 0,
            'total_procesados': 0,
            'total_emails': 0
        }
